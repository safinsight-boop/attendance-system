#!/usr/bin/env python3
"""
Attendance Management System — TTLock Integration
Flask + SQLite + APScheduler + Gmail SMTP
"""
import os, sqlite3, hashlib, requests, smtplib, logging, io, secrets, threading
from datetime import datetime, timedelta, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, request, jsonify, render_template, send_file, session
from functools import wraps
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-change-in-prod')
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════
DB_PATH    = os.environ.get('DB_PATH', '/data/attendance.db')
SUPABASE_URL = os.getenv('SUPABASE_URL', 'https://nxraodhjulwsmldjtyyv.supabase.co')
SUPABASE_KEY = os.getenv('SUPABASE_KEY', '')
TTBASE     = os.getenv('TTLOCK_BASE_URL', 'https://euapi.ttlock.com')
CID        = os.getenv('TTLOCK_CLIENT_ID', '')
CSECRET    = os.getenv('TTLOCK_CLIENT_SECRET', '')
TTUSR      = os.getenv('TTLOCK_USERNAME', '')
TTPASS     = os.getenv('TTLOCK_PASSWORD', '')

def _tt_creds():
    """Return (cid, csecret, usr, pwd) — DB settings take priority, then env vars."""
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT key, value FROM settings WHERE key IN "
            "('tt_client_id','tt_client_secret','tt_username','tt_password')"
        ).fetchall()
        conn.close()
        s = {r['key']: r['value'] for r in rows}
        db_cid  = s.get('tt_client_id','')
        db_csec = s.get('tt_client_secret','')
        db_usr  = s.get('tt_username','')
        db_pwd  = s.get('tt_password','')
        if db_cid and db_usr and db_pwd:
            return db_cid, db_csec, db_usr, db_pwd
    except Exception:
        pass
    return CID, CSECRET, TTUSR, TTPASS
EMAIL_FROM       = os.getenv('EMAIL_SENDER', '')
EMAIL_PASS       = os.getenv('EMAIL_PASSWORD', '')
SITE_URL         = os.getenv('SITE_URL', 'https://attendance-system-pd27.onrender.com')
GRACE_MIN        = 5    # grace minutes before marking late

# ═══════════════════════════════════════════════════════════
#  SUPABASE AUDIT LOG
# ═══════════════════════════════════════════════════════════
def audit_log(event_type, target_type='', target_name='', details='', status='success'):
    """Log events to Supabase — sent in a background thread to avoid slowing the app"""
    if not SUPABASE_KEY:
        return
    username = session.get('username', 'system')
    role     = session.get('role', '')
    ip       = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    payload  = {
        'event_type':  event_type,
        'username':    username,
        'role':        role,
        'target_type': target_type,
        'target_name': target_name,
        'details':     details,
        'ip_address':  ip.split(',')[0].strip(),
        'status':      status,
    }
    def _send():
        try:
            requests.post(
                f"{SUPABASE_URL}/rest/v1/audit_logs",
                json=payload,
                headers={
                    'apikey':        SUPABASE_KEY,
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'Content-Type':  'application/json',
                    'Prefer':        'return=minimal',
                },
                timeout=5
            )
        except Exception as e:
            logger.warning(f"audit_log failed: {e}")
    threading.Thread(target=_send, daemon=True).start()
AUTO_REJECT_DAYS = 3    # days until automatic rejection of pending requests

# Violations table: {bracket: [(ptype, pvalue), ...]}
# ptype: 'warning' | 'percent' | 'day' | 'warning_day'
# percent → % of daily wage | day → N × daily wage
PENALTIES = {
    # Late arrival — without disrupting other workers (Labor Regulations)
    'late_1_15':   [('warning', 0),  ('percent', 5),  ('percent', 10), ('percent', 20)],  # ≤15 min
    'late_15_30':  [('percent', 10), ('percent', 25), ('percent', 50), ('day', 1)],        # 15-30 min
    'late_30_60':  [('percent', 25), ('percent', 50), ('percent', 75), ('day', 1)],        # 30-60 min
    'late_60plus': [('warning', 0),  ('day', 1),      ('day', 2),      ('day', 3)],        # >60 min
    # Early departure
    'early_u15':   [('warning', 0),  ('percent', 10), ('percent', 25), ('day', 1)],        # ≤15 min
    'early_o15':   [('percent', 10), ('percent', 25), ('percent', 50), ('day', 1)],        # >15 min
    # Flexible schedule employees
    'flex_hours':  [('hours', 0)],
    # Absence without excuse
    'absent_1':    [('percent', 50), ('day', 1), ('day', 2), ('day', 3)],
}

MONTHS_AR = ['January','February','March','April','May','June',
             'July','August','September','October','November','December']

_tt_cache = {'token': None, 'exp': 0}

# ═══════════════════════════════════════════════════════════
#  DATABASE
# ═══════════════════════════════════════════════════════════
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    conn = get_db()
    try:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS employees (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            name_ar      TEXT NOT NULL,
            name_en      TEXT NOT NULL UNIQUE,
            email        TEXT,
            salary       REAL DEFAULT 0,
            housing      REAL DEFAULT 0,
            transport    REAL DEFAULT 0,
            commission   REAL DEFAULT 0,
            other_ded    REAL DEFAULT 0,
            work_type    TEXT DEFAULT 'fixed',
            work_start   TEXT DEFAULT '08:00',
            work_end     TEXT DEFAULT '17:00',
            weekly_hours      REAL DEFAULT 40,
            annual_leave_days INTEGER DEFAULT 21,
            emp_code          TEXT,
            created_at        TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS attendance (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            att_date    TEXT NOT NULL,
            check_in    TEXT,
            check_out   TEXT,
            late_min    INTEGER DEFAULT 0,
            early_min   INTEGER DEFAULT 0,
            total_hours REAL DEFAULT 0,
            status      TEXT DEFAULT 'present',
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
            UNIQUE(employee_id, att_date)
        );

        CREATE TABLE IF NOT EXISTS violations (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            vio_date    TEXT NOT NULL,
            vtype       TEXT NOT NULL,
            bracket     TEXT NOT NULL,
            occurrence  INTEGER NOT NULL,
            ptype       TEXT NOT NULL,
            pvalue      REAL DEFAULT 0,
            deduction   REAL DEFAULT 0,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS vio_counts (
            employee_id INTEGER NOT NULL,
            yr          INTEGER NOT NULL,
            mo          INTEGER NOT NULL,
            bracket     TEXT NOT NULL,
            cnt         INTEGER DEFAULT 0,
            PRIMARY KEY (employee_id, yr, mo, bracket)
        );

        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role          TEXT NOT NULL DEFAULT 'employee',
            employee_id   INTEGER,
            reset_token   TEXT,
            reset_exp     TEXT,
            created_at    TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS excuse_requests (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id  INTEGER NOT NULL,
            att_date     TEXT NOT NULL,
            vtype        TEXT NOT NULL,
            reason       TEXT NOT NULL,
            submitted_at TEXT DEFAULT (datetime('now')),
            status       TEXT DEFAULT 'pending',
            decided_by   INTEGER,
            decided_at   TEXT,
            manager_note TEXT,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS leaves (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id  INTEGER NOT NULL,
            leave_type   TEXT NOT NULL,
            start_date   TEXT NOT NULL,
            end_date     TEXT NOT NULL,
            days         INTEGER NOT NULL,
            status       TEXT DEFAULT 'pending',
            approved_by  INTEGER,
            sick_doc     TEXT,
            notes        TEXT,
            created_at   TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS public_holidays (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            h_date     TEXT UNIQUE NOT NULL,
            name       TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS overtime_requests (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id    INTEGER NOT NULL,
            att_date       TEXT NOT NULL,
            overtime_hours REAL NOT NULL,
            check_out_time TEXT NOT NULL,
            work_end       TEXT NOT NULL,
            status         TEXT DEFAULT 'pending',
            decided_by     INTEGER,
            decided_at     TEXT,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS attendance_requests (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id     INTEGER NOT NULL,
            req_date        TEXT NOT NULL,
            req_type        TEXT NOT NULL,
            reason          TEXT NOT NULL,
            requested_time  TEXT,
            attachment      TEXT,
            attachment_name TEXT,
            submitted_at    TEXT DEFAULT (datetime('now')),
            status          TEXT DEFAULT 'pending',
            decided_by      INTEGER,
            decided_at      TEXT,
            manager_note    TEXT,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        INSERT OR IGNORE INTO settings VALUES ('company_name',    'Employee Portal');
        INSERT OR IGNORE INTO settings VALUES ('company_name_en', 'Employee Portal');
        INSERT OR IGNORE INTO settings VALUES ('weekend_days',    '5,6');
        """)
        conn.commit()
        _migrate_db(conn)
        _seed_default_user(conn)
        _seed_ttlock_from_env(conn)
        logger.info("Database initialized OK")
    finally:
        conn.close()

def _migrate_db(conn):
    """Add new columns to existing tables (safe to run repeatedly)"""
    migrations = [
        "ALTER TABLE excuse_requests ADD COLUMN attachment TEXT",
        "ALTER TABLE excuse_requests ADD COLUMN attachment_name TEXT",
        "ALTER TABLE leaves ADD COLUMN attachment TEXT",
        "ALTER TABLE leaves ADD COLUMN attachment_name TEXT",
        "ALTER TABLE employees ADD COLUMN annual_leave_days INTEGER DEFAULT 21",
        "ALTER TABLE employees ADD COLUMN emp_code TEXT",
        "ALTER TABLE overtime_requests ADD COLUMN notes TEXT",
        "ALTER TABLE overtime_requests ADD COLUMN source TEXT DEFAULT 'auto'",
        "ALTER TABLE employees ADD COLUMN weekend_days TEXT DEFAULT '5,6'",
        "ALTER TABLE overtime_requests ADD COLUMN manager_note TEXT",
        "ALTER TABLE employees ADD COLUMN nationality TEXT",
        "ALTER TABLE employees ADD COLUMN department TEXT",
        "ALTER TABLE employees ADD COLUMN job_title TEXT",
        "ALTER TABLE employees ADD COLUMN status TEXT DEFAULT 'active'",
        "ALTER TABLE employees ADD COLUMN hire_date TEXT",
        "ALTER TABLE employees ADD COLUMN national_id TEXT",
        "ALTER TABLE employees ADD COLUMN phone TEXT",
        "ALTER TABLE employees ADD COLUMN direct_manager TEXT",
        "ALTER TABLE employees ADD COLUMN notes TEXT",
        "ALTER TABLE employees ADD COLUMN contract_type TEXT DEFAULT 'permanent'",
        "ALTER TABLE employees ADD COLUMN contract_start TEXT",
        "ALTER TABLE employees ADD COLUMN contract_end TEXT",
        "ALTER TABLE employees ADD COLUMN probation_end TEXT",
        "ALTER TABLE employees ADD COLUMN iqama_expiry TEXT",
        "ALTER TABLE employees ADD COLUMN gosi_number TEXT",
        "ALTER TABLE employees ADD COLUMN iban TEXT",
        "ALTER TABLE employees ADD COLUMN medical_insurance TEXT",
        """CREATE TABLE IF NOT EXISTS schedule_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id  INTEGER NOT NULL,
            effective_date TEXT NOT NULL,
            work_type    TEXT,
            work_start   TEXT,
            work_end     TEXT,
            weekly_hours REAL,
            weekend_days TEXT,
            recorded_at  TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        )""",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
        except Exception:
            pass
    conn.commit()

def _seed_ttlock_from_env(conn):
    """Seed TTLock credentials from env vars into DB if DB is empty — env vars win on fresh DB."""
    if not CID or not TTUSR or not TTPASS:
        return
    existing = conn.execute(
        "SELECT value FROM settings WHERE key='tt_client_id'"
    ).fetchone()
    if existing and existing['value']:
        return  # DB already has credentials — don't overwrite
    for key, val in [('tt_client_id', CID), ('tt_client_secret', CSECRET),
                     ('tt_username', TTUSR), ('tt_password', TTPASS)]:
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (key, val))
    conn.commit()
    logger.info("TTLock credentials seeded from environment variables")

def _seed_default_user(conn):
    count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count == 0:
        pw = hashlib.sha256('admin123'.encode()).hexdigest()
        conn.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
            ('admin', pw, 'hr'))
        conn.commit()
        logger.info("Default HR user created — username: admin / password: admin123")

# ═══════════════════════════════════════════════════════════
#  AUTH HELPERS
# ═══════════════════════════════════════════════════════════
def _hash(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized', 'login_required': True}), 401
        return f(*args, **kwargs)
    return decorated

def hr_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized', 'login_required': True}), 401
        if session.get('role') not in ('hr', 'manager'):
            return jsonify({'error': 'Insufficient permissions'}), 403
        return f(*args, **kwargs)
    return decorated

def _gosi(emp):
    """Calculate GOSI insurance = 10.75% — trainees (emp_code starting with IN) are exempt"""
    code = (emp.get('emp_code') or '').strip().lower()
    if code.startswith('in'):
        return 0.0
    return round(((emp['salary'] or 0) + (emp['housing'] or 0) + (emp['transport'] or 0)) * 0.1075, 2)

def _years_of_service(hire_date_str):
    if not hire_date_str:
        return None
    try:
        hire = date.fromisoformat(hire_date_str)
        return round((date.today() - hire).days / 365.25, 1)
    except Exception:
        return None

def _eosb(salary, years):
    if not years or years < 2:
        return 0.0
    first5 = min(years, 5) * 0.5 * (salary or 0)
    after5 = max(0.0, years - 5) * (salary or 0)
    return round(first5 + after5, 2)

def _iqama_alert(iqama_expiry_str):
    if not iqama_expiry_str:
        return None
    try:
        exp = date.fromisoformat(iqama_expiry_str)
        days_left = (exp - date.today()).days
        if days_left < 0:
            return f"إقامة منتهية منذ {abs(days_left)} يوم"
        if days_left <= 60:
            return f"⚠ إقامة تنتهي خلال {days_left} يوم"
        return None
    except Exception:
        return None

def _leave_balance(conn, emp_id, year=None):
    """Annual leave balance: entitlement - used = remaining"""
    if year is None:
        year = date.today().year
    emp = conn.execute("SELECT annual_leave_days FROM employees WHERE id=?", (emp_id,)).fetchone()
    entitlement = (emp['annual_leave_days'] if emp and emp['annual_leave_days'] else 21)
    used = conn.execute("""
        SELECT COALESCE(SUM(days), 0) AS d FROM leaves
        WHERE employee_id=? AND leave_type='annual' AND status='approved'
          AND strftime('%Y', start_date) = ?
    """, (emp_id, str(year))).fetchone()['d'] or 0
    return {
        'entitlement': entitlement,
        'used':        int(used),
        'remaining':   max(0, entitlement - int(used)),
        'year':        year,
    }

def _is_on_leave(conn, emp_id, target_date):
    """Is the employee on approved leave on this date?"""
    ds = str(target_date)
    # public holiday
    if conn.execute("SELECT 1 FROM public_holidays WHERE h_date=?", (ds,)).fetchone():
        return 'official_holiday'
    # approved personal leave
    row = conn.execute("""
        SELECT leave_type FROM leaves
        WHERE employee_id=? AND status='approved'
        AND start_date<=? AND end_date>=?
    """, (emp_id, ds, ds)).fetchone()
    return row['leave_type'] if row else None

# ═══════════════════════════════════════════════════════════
#  TTLOCK API
# ═══════════════════════════════════════════════════════════
def _md5(s):
    return hashlib.md5(s.encode()).hexdigest()

def tt_get_token(return_error=False):
    global _tt_cache
    now = datetime.now().timestamp()
    if not return_error and _tt_cache['token'] and now < _tt_cache['exp'] - 120:
        return _tt_cache['token']
    cid, csecret, usr, pwd = _tt_creds()
    if not cid or not usr:
        return (None, 'missing_credentials') if return_error else None
    try:
        r = requests.post(f"{TTBASE}/oauth2/token", data={
            'client_id':     cid,
            'client_secret': csecret,
            'grant_type':    'password',
            'username':      usr,
            'password':      _md5(pwd),
        }, timeout=15)
        d = r.json()
        if 'access_token' in d:
            _tt_cache = {'token': d['access_token'], 'exp': now + d.get('expires_in', 7200)}
            logger.info("TTLock token refreshed")
            return (d['access_token'], None) if return_error else d['access_token']
        logger.error(f"TTLock auth failed: {d}")
        return (None, d) if return_error else None
    except Exception as e:
        logger.error(f"TTLock auth error: {e}")
        return (None, str(e)) if return_error else None

def tt_get_locks(token):
    locks, page = [], 1
    ts = int(datetime.now().timestamp() * 1000)
    cid = _tt_creds()[0]
    while True:
        try:
            r = requests.get(f"{TTBASE}/v3/lock/list", params={
                'clientId': cid, 'accessToken': token,
                'pageNo': page, 'pageSize': 100, 'date': ts
            }, timeout=15).json()
            if r.get('errcode', 0) != 0:
                logger.warning(f"tt_get_locks err: {r}")
                break
            batch = r.get('list', [])
            locks.extend(batch)
            if len(batch) < 100: break
            page += 1
        except Exception as e:
            logger.error(f"tt_get_locks: {e}"); break
    return locks

def tt_get_records(token, lock_id, start_ms, end_ms):
    recs, page = [], 1
    ts = int(datetime.now().timestamp() * 1000)
    cid = _tt_creds()[0]
    while True:
        try:
            r = requests.get(f"{TTBASE}/v3/lockRecord/list", params={
                'clientId': cid, 'accessToken': token,
                'lockId': lock_id, 'startDate': start_ms,
                'endDate': end_ms, 'pageNo': page,
                'pageSize': 100, 'date': ts
            }, timeout=15).json()
            if r.get('errcode', 0) != 0: break
            batch = r.get('list', [])
            recs.extend(batch)
            if len(batch) < 100: break
            page += 1
        except Exception as e:
            logger.error(f"tt_get_records lock={lock_id}: {e}"); break
    return recs

def fetch_daily_records(target_date):
    """
    Fetch TTLock records for a specific day (Riyadh timezone UTC+3).
    Returns: {name_en_lower: [datetime, ...]} sorted chronologically
    """
    token = tt_get_token()
    if not token:
        logger.error("No TTLock token — cannot fetch records")
        return None

    TZ_OFFSET = 3 * 3600  # Riyadh = UTC+3
    day_start = int(datetime(
        target_date.year, target_date.month, target_date.day, 0, 0, 0
    ).timestamp() * 1000) - TZ_OFFSET * 1000
    day_end = int(datetime(
        target_date.year, target_date.month, target_date.day, 23, 59, 59
    ).timestamp() * 1000) - TZ_OFFSET * 1000

    by_user = {}
    for lock in tt_get_locks(token):
        lid = lock.get('lockId')
        if not lid: continue
        for rec in tt_get_records(token, lid, day_start, day_end):
            if not rec.get('success'): continue
            uname = (rec.get('username') or '').strip().lower()
            ts_ms = rec.get('lockDate') or rec.get('serverDate', 0)
            if uname and ts_ms:
                riyadh_time = datetime.fromtimestamp(ts_ms / 1000) + timedelta(hours=3)
                by_user.setdefault(uname, []).append(riyadh_time)

    for k in by_user:
        by_user[k].sort()
    logger.info(f"TTLock: fetched records for {len(by_user)} users on {target_date}")
    return by_user

# ═══════════════════════════════════════════════════════════
#  VIOLATION ENGINE
# ═══════════════════════════════════════════════════════════
def late_bracket(mins):
    if mins <= 0:  return None
    if mins <= 15: return 'late_1_15'
    if mins <= 30: return 'late_15_30'
    if mins <= 60: return 'late_30_60'
    return 'late_60plus'

def early_bracket(mins):
    if mins <= 0:   return None
    if mins <= 15:  return 'early_u15'   # ≤15 min (inclusive)
    return 'early_o15'                   # >15 min

def next_occurrence(conn, emp_id, yr, mo, bracket):
    """Increment violation counter and return new count"""
    row = conn.execute(
        "SELECT cnt FROM vio_counts WHERE employee_id=? AND yr=? AND mo=? AND bracket=?",
        (emp_id, yr, mo, bracket)
    ).fetchone()
    cnt = (row['cnt'] if row else 0) + 1
    conn.execute("""
        INSERT INTO vio_counts (employee_id, yr, mo, bracket, cnt)
        VALUES (?,?,?,?,?)
        ON CONFLICT(employee_id,yr,mo,bracket) DO UPDATE SET cnt=excluded.cnt
    """, (emp_id, yr, mo, bracket, cnt))
    return cnt

def calc_deduction(emp, ptype, pvalue):
    """Calculate monetary deduction amount"""
    daily = (emp['salary'] + emp['housing'] + emp['transport']) / 30
    if ptype == 'warning':               return 0.0
    if ptype == 'percent':               return round(daily * pvalue / 100, 2)
    if ptype in ('day', 'warning_day'):  return round(daily * pvalue, 2)
    return 0.0

def _get_emp_schedule(conn, emp, date_str):
    """Returns effective schedule data for a given date (checks effective date)."""
    row = conn.execute("""
        SELECT * FROM schedule_history
        WHERE employee_id=? AND effective_date<=?
        ORDER BY effective_date DESC LIMIT 1
    """, (emp['id'], date_str)).fetchone()
    if row:
        merged = dict(emp)
        merged.update({
            'work_type':    row['work_type']    or emp['work_type'],
            'work_start':   row['work_start']   or emp['work_start'],
            'work_end':     row['work_end']      or emp['work_end'],
            'weekly_hours': row['weekly_hours']  or emp['weekly_hours'],
            'weekend_days': row['weekend_days']  or emp.get('weekend_days','5,6'),
        })
        return merged
    return emp

def apply_violation(conn, emp, vio_date, vtype, bracket):
    """Apply violation and save it, returns (ptype, pvalue, deduction)"""
    yr, mo = vio_date.year, vio_date.month
    occ = next_occurrence(conn, emp['id'], yr, mo, bracket)
    idx = min(occ, 4) - 1
    ptype, pvalue = PENALTIES[bracket][idx]
    ded = calc_deduction(emp, ptype, pvalue)
    conn.execute("""
        INSERT INTO violations
            (employee_id, vio_date, vtype, bracket, occurrence, ptype, pvalue, deduction)
        VALUES (?,?,?,?,?,?,?,?)
    """, (emp['id'], str(vio_date), vtype, bracket, occ, ptype, pvalue, ded))
    return ptype, pvalue, ded

# ═══════════════════════════════════════════════════════════
#  EMAIL NOTIFICATIONS
# ═══════════════════════════════════════════════════════════
def send_email(to, subject, html_body):
    if not EMAIL_FROM or not EMAIL_PASS or not to:
        return
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = EMAIL_FROM
        msg['To']      = to
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_FROM, EMAIL_PASS)
            smtp.sendmail(EMAIL_FROM, to, msg.as_bytes())
        logger.info(f"Email sent → {to}: {subject}")
    except Exception as e:
        logger.error(f"Email error: {e}")

_STYLE = "font-family:Tahoma,Arial;direction:ltr;padding:28px;max-width:620px;margin:0 auto;color:#1e293b"
_TABLE = "border-collapse:collapse;width:100%;margin-top:14px"
_TD    = "padding:10px 14px;border:1px solid #e2e8f0;font-size:14px"

def notify_attendance(emp, att_date, status, check_in, check_out,
                      late_min, early_min, ptype=None, pvalue=None, ded=0.0):
    if not emp.get('email'): return
    name = emp.get('name_ar') or emp.get('name_en', '')
    ds   = str(att_date)
    ci   = check_in.strftime('%I:%M %p')  if check_in  else '—'
    co   = check_out.strftime('%I:%M %p') if check_out else '—'

    def penalty_text():
        if not ptype or ptype == 'warning': return 'Written Warning'
        if ptype == 'percent': return f"Deduction: {pvalue}% of daily wage"
        if ptype in ('day', 'warning_day'): return f"Deduction: {pvalue} day(s) from salary"
        return '—'

    if status == 'on_time':
        subj = f"✅ Attendance Commendation — {ds}"
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#16a34a;margin-bottom:6px">✅ Commendation</h2>
          <p>Dear <b>{name}</b>,</p>
          <p>Thank you for your punctuality today <b>{ds}</b>.</p>
          <table style="{_TABLE}">
            <tr><td style="{_TD};background:#f8fafc">Check-in</td><td style="{_TD}"><b>{ci}</b></td></tr>
            <tr><td style="{_TD};background:#f8fafc">Check-out</td><td style="{_TD}"><b>{co}</b></td></tr>
          </table>
          <p style="color:#64748b;margin-top:16px;font-size:13px">We appreciate your dedication and discipline.</p></div>"""

    elif status == 'late':
        subj = f"⚠️ Late Arrival Notice — {ds}"
        ded_row = f'<tr><td style="{_TD};background:#fff1f2">Deduction Amount</td><td style="{_TD};color:#dc2626"><b>{ded:.2f} SR</b></td></tr>' if ded > 0 else ''
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#d97706;margin-bottom:6px">⚠️ Violation Notice — Late Arrival</h2>
          <p>Dear <b>{name}</b>,</p>
          <p>A late arrival was recorded on <b>{ds}</b>.</p>
          <table style="{_TABLE}">
            <tr><td style="{_TD};background:#f8fafc">Check-in Time</td><td style="{_TD}"><b>{ci}</b></td></tr>
            <tr><td style="{_TD};background:#fff7ed">Late Duration</td><td style="{_TD};color:#dc2626"><b>{late_min} min</b></td></tr>
            <tr><td style="{_TD};background:#f8fafc">Penalty</td><td style="{_TD}"><b>{penalty_text()}</b></td></tr>
            {ded_row}
          </table>
          <p style="margin-top:16px">If you have a valid excuse, you can submit it directly:</p>
          <a href="{SITE_URL}" style="display:inline-block;background:#3b82f6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:700;margin-top:6px">
            Submit Excuse
          </a>
          <p style="color:#94a3b8;font-size:12px;margin-top:12px">The excuse will be automatically rejected if not submitted within {AUTO_REJECT_DAYS} days.</p>
        </div>"""

    elif status == 'early_leave':
        subj = f"⚠️ Early Departure Notice — {ds}"
        ded_row = f'<tr><td style="{_TD};background:#fff1f2">Deduction Amount</td><td style="{_TD};color:#dc2626"><b>{ded:.2f} SR</b></td></tr>' if ded > 0 else ''
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#d97706;margin-bottom:6px">⚠️ Violation Notice — Early Departure</h2>
          <p>Dear <b>{name}</b>,</p>
          <p>An early departure was recorded on <b>{ds}</b>.</p>
          <table style="{_TABLE}">
            <tr><td style="{_TD};background:#f8fafc">Check-out Time</td><td style="{_TD}"><b>{co}</b></td></tr>
            <tr><td style="{_TD};background:#fff7ed">Early Departure</td><td style="{_TD};color:#dc2626"><b>{early_min} min</b></td></tr>
            <tr><td style="{_TD};background:#f8fafc">Penalty</td><td style="{_TD}"><b>{penalty_text()}</b></td></tr>
            {ded_row}
          </table>
          <p style="margin-top:16px">If you have a valid excuse, you can submit it directly:</p>
          <a href="{SITE_URL}" style="display:inline-block;background:#3b82f6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:700;margin-top:6px">
            Submit Excuse
          </a>
          <p style="color:#94a3b8;font-size:12px;margin-top:12px">The excuse will be automatically rejected if not submitted within {AUTO_REJECT_DAYS} days.</p>
        </div>"""

    elif status == 'absent':
        subj = f"🔴 Absence Notice — {ds}"
        ded_row = f'<tr><td style="{_TD};background:#fff1f2">Deduction Amount</td><td style="{_TD};color:#dc2626"><b>{ded:.2f} SR</b></td></tr>' if ded > 0 else ''
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#dc2626;margin-bottom:6px">🔴 Absence Notice</h2>
          <p>Dear <b>{name}</b>,</p>
          <p>No attendance was recorded for you on <b>{ds}</b>.</p>
          <table style="{_TABLE}">
            <tr><td style="{_TD};background:#f8fafc">Penalty</td><td style="{_TD}"><b>{penalty_text()}</b></td></tr>
            {ded_row}
          </table>
          <p style="margin-top:16px">If you have a valid excuse, you can submit it directly:</p>
          <a href="{SITE_URL}" style="display:inline-block;background:#3b82f6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:700;margin-top:6px">
            Submit Excuse
          </a>
          <p style="color:#94a3b8;font-size:12px;margin-top:12px">The excuse will be automatically rejected if not submitted within {AUTO_REJECT_DAYS} days.</p>
        </div>"""
    elif status.startswith('leave_'):
        return  # no notification for approved leaves
    else:
        return

    send_email(emp['email'], subj, body)

def _notify_overtime(emp, att_date, ot_hours, checkout_time, conn):
    """Notify HR managers and department of unconfirmed overtime"""
    managers = conn.execute(
        "SELECT u.*, e.email AS emp_email FROM users u "
        "LEFT JOIN employees e ON e.id=u.employee_id "
        "WHERE u.role IN ('hr','manager')"
    ).fetchall()
    name = emp.get('name_ar') or emp.get('name_en', '')
    for mgr in managers:
        to = mgr['emp_email'] or EMAIL_FROM
        if not to:
            continue
        subj = f"🕐 Unconfirmed Overtime — {name} — {att_date}"
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#d97706;margin-bottom:6px">🕐 Overtime Notice</h2>
          <p>Employee <b>{name}</b> stayed <b>{ot_hours:.1f} hr</b> beyond their shift on <b>{att_date}</b>.</p>
          <p>Check-out time: <b>{checkout_time}</b></p>
          <p>Is there an <b>official assignment</b> for this overtime?</p>
          <p>
            <a href="{SITE_URL}" style="background:#3b82f6;color:#fff;padding:10px 20px;border-radius:8px;text-decoration:none;font-weight:700">
              Review on Website
            </a>
          </p>
          <p style="color:#94a3b8;font-size:12px;margin-top:14px">If not confirmed within 24 hours, the overtime will not be counted.</p>
        </div>"""
        send_email(to, subj, body)

def notify_flex_weekly(emp, friday, actual_h, required_h, ded):
    if not emp.get('email'): return
    name = emp.get('name_ar') or emp.get('name_en', '')
    monday = friday - timedelta(days=4)
    missing = max(0.0, required_h - actual_h)
    color = "#dc2626" if missing > 0 else "#16a34a"
    subj = f"📊 Weekly Hours Report — {monday} to {friday}"
    ded_row = f'<tr><td style="{_TD};background:#fff1f2;color:#dc2626">Estimated Deduction</td><td style="{_TD};color:#dc2626"><b>{ded:.2f} SR</b></td></tr>' if ded > 0 else ''
    body = f"""<div style="{_STYLE}">
      <h2 style="margin-bottom:6px">📊 Weekly Hours Report</h2>
      <p>Dear <b>{name}</b>,</p>
      <p>Week: <b>{monday}</b> — <b>{friday}</b></p>
      <table style="{_TABLE}">
        <tr><td style="{_TD};background:#f8fafc">Required Hours</td><td style="{_TD}"><b>{required_h:.1f} hr</b></td></tr>
        <tr><td style="{_TD};background:#f8fafc">Actual Hours</td><td style="{_TD}"><b>{actual_h:.1f} hr</b></td></tr>
        <tr><td style="{_TD};background:#f8fafc;color:{color}">Missing Hours</td><td style="{_TD};color:{color}"><b>{missing:.1f} hr</b></td></tr>
        {ded_row}
      </table></div>"""
    send_email(emp['email'], subj, body)

# ═══════════════════════════════════════════════════════════
#  ATTENDANCE PROCESSING ENGINE
# ═══════════════════════════════════════════════════════════
def process_day(target_date=None):
    if target_date is None:
        target_date = date.today()

    # skip attendance processing if TTLock is not configured
    if not CID or not TTUSR:
        logger.warning("TTLock not configured — skipping attendance processing")
        return {'ok': False, 'msg': 'TTLock not connected — cannot process attendance'}

    logger.info(f"=== Processing attendance for {target_date} ===")
    raw = fetch_daily_records(target_date)

    # if fetching records failed (token failed) do not auto-record absence
    if raw is None:
        logger.warning(f"fetch_daily_records returned None for {target_date} — skipping")
        return {'ok': False, 'msg': 'Failed to connect to TTLock'}

    conn = get_db()
    try:
        employees = conn.execute("SELECT * FROM employees").fetchall()

        for emp_row in employees:
            emp = _get_emp_schedule(conn, dict(emp_row), str(target_date))
            uname = emp['name_en'].strip().lower()
            times = raw.get(uname, [])

            check_in  = times[0]  if times          else None
            check_out = times[-1] if len(times) > 1 else None

            late_min = early_min = 0
            total_hours = 0.0
            status = 'absent'
            ptype = pvalue = None
            ded = 0.0

            if not times:
                # check leave before recording absence
                leave_type = _is_on_leave(conn, emp['id'], target_date)
                if leave_type:
                    status = f'leave_{leave_type}'
                else:
                    status = 'absent'
                    # single-day absence violation
                    ptype, pvalue, ded = apply_violation(
                        conn, emp, target_date, 'absent', 'absent_1')

            elif emp['work_type'] == 'fixed':
                try:
                    wstart = datetime.strptime(
                        f"{target_date} {emp['work_start']}", "%Y-%m-%d %H:%M")
                    wend   = datetime.strptime(
                        f"{target_date} {emp['work_end']}",   "%Y-%m-%d %H:%M")
                except Exception:
                    wstart = wend = None

                if check_in and wstart:
                    raw_late = (check_in - wstart).total_seconds() / 60
                    # 5-minute grace period
                    late_min = max(0, int(raw_late) - GRACE_MIN)

                if check_out and wend:
                    diff = (wend - check_out).total_seconds() / 60
                    early_min = max(0, int(diff))

                if check_in and check_out:
                    total_hours = round(
                        (check_out - check_in).total_seconds() / 3600, 2)

                # -- priority: late first, then early departure --
                if late_min > 0:
                    status = 'late'
                    br = late_bracket(late_min)
                    ptype, pvalue, ded = apply_violation(
                        conn, emp, target_date, 'late', br)
                    if early_min > 0:
                        apply_violation(conn, emp, target_date,
                                        'early_leave', early_bracket(early_min))
                elif early_min > 0:
                    status = 'early_leave'
                    br = early_bracket(early_min)
                    ptype, pvalue, ded = apply_violation(
                        conn, emp, target_date, 'early_leave', br)
                else:
                    status = 'on_time'

                # -- detect overtime (>30 min after scheduled end) --
                if check_out and wend:
                    ot_min = (check_out - wend).total_seconds() / 60
                    if ot_min > 30:
                        ot_hours = round(ot_min / 60, 2)
                        conn.execute("""
                            INSERT OR IGNORE INTO overtime_requests
                                (employee_id, att_date, overtime_hours, check_out_time, work_end)
                            VALUES (?,?,?,?,?)
                        """, (emp['id'], str(target_date), ot_hours,
                              check_out.strftime('%H:%M'), emp['work_end']))
                        _notify_overtime(emp, target_date, ot_hours,
                                         check_out.strftime('%H:%M'), conn)

            else:  # flex
                if check_in and check_out:
                    total_hours = round(
                        (check_out - check_in).total_seconds() / 3600, 2)
                status = 'present'

            # save attendance record
            conn.execute("""
                INSERT INTO attendance
                    (employee_id, att_date, check_in, check_out,
                     late_min, early_min, total_hours, status)
                VALUES (?,?,?,?,?,?,?,?)
                ON CONFLICT(employee_id, att_date) DO UPDATE SET
                    check_in=excluded.check_in,
                    check_out=excluded.check_out,
                    late_min=excluded.late_min,
                    early_min=excluded.early_min,
                    total_hours=excluded.total_hours,
                    status=excluded.status
            """, (
                emp['id'], str(target_date),
                check_in.strftime('%H:%M')  if check_in  else None,
                check_out.strftime('%H:%M') if check_out else None,
                late_min, early_min, total_hours, status
            ))

            # send notification email
            notify_attendance(emp, target_date, status, check_in, check_out,
                              late_min, early_min, ptype, pvalue, ded)

        # -- check flex employees (every Friday) --
        if target_date.weekday() == 4:
            monday = target_date - timedelta(days=4)
            week_dates = [str(monday + timedelta(days=i)) for i in range(5)]
            ph = ','.join(['?'] * 5)

            flex_emps = conn.execute(
                "SELECT * FROM employees WHERE work_type='flex'"
            ).fetchall()

            for emp_row in flex_emps:
                emp = dict(emp_row)
                row = conn.execute(
                    f"SELECT COALESCE(SUM(total_hours),0) AS h FROM attendance "
                    f"WHERE employee_id=? AND att_date IN ({ph})",
                    [emp['id']] + week_dates
                ).fetchone()

                actual_h   = row['h'] or 0.0
                required_h = emp['weekly_hours'] or 40.0
                missing    = max(0.0, required_h - actual_h)

                if missing > 0:
                    hourly = (emp['salary'] + emp['housing'] + emp['transport']
                              ) / (4 * required_h)
                    flex_ded = round(missing * hourly, 2)
                    occ = next_occurrence(
                        conn, emp['id'],
                        target_date.year, target_date.month, 'flex_hours')
                    conn.execute("""
                        INSERT INTO violations
                            (employee_id,vio_date,vtype,bracket,occurrence,ptype,pvalue,deduction)
                        VALUES (?,?,?,?,?,?,?,?)
                    """, (emp['id'], str(target_date), 'flex_hours',
                          'flex_hours', occ, 'hours', missing, flex_ded))
                    notify_flex_weekly(emp, target_date, actual_h, required_h, flex_ded)
                else:
                    notify_flex_weekly(emp, target_date, actual_h, required_h, 0.0)

        conn.commit()
        logger.info(f"=== Attendance processing done: {target_date} ===")

    except Exception as e:
        conn.rollback()
        logger.error(f"process_day error: {e}", exc_info=True)
        raise
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════════
#  EXCEL — HELPERS
# ═══════════════════════════════════════════════════════════
def _bdr():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bold=False, bg=None, fg='000000',
          size=11, wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=size, color=fg,
                       name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center',
                            wrap_text=wrap, reading_order=2)
    c.border    = _bdr()
    if bg:      c.fill = PatternFill('solid', fgColor=bg)
    if num_fmt: c.number_format = num_fmt
    return c

# ═══════════════════════════════════════════════════════════
#  EXCEL — ATTENDANCE EXPORT
# ═══════════════════════════════════════════════════════════
STATUS_AR = {
    'on_time': 'On Time', 'late': 'Late',
    'absent': 'Absent', 'early_leave': 'Early Departure', 'present': 'Present'
}

def export_attendance_excel(year, month, emp_id=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Log"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    col_widths = [22, 12, 10, 10, 10, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 24

    # Title
    ws.merge_cells('A1:G1')
    _cell(ws, 1, 1,
          f"Attendance Log — {MONTHS_AR[month-1]} {year}",
          bold=True, size=15, bg='1F4E79', fg='FFFFFF')

    # Headers
    hdrs = ['Employee Name', 'Date', 'Check-in', 'Check-out',
            'Late (min)', 'Early Exit (min)', 'Status']
    for i, h in enumerate(hdrs, 1):
        _cell(ws, 2, i, h, bold=True, bg='2E74B5', fg='FFFFFF')

    conn = get_db()
    try:
        q = """SELECT e.name_ar, a.*
               FROM attendance a
               JOIN employees e ON e.id = a.employee_id
               WHERE a.att_date LIKE ?"""
        params = [f"{year}-{month:02d}-%"]
        if emp_id:
            q += " AND a.employee_id=?"
            params.append(emp_id)
        q += " ORDER BY a.att_date, e.name_ar"
        rows = conn.execute(q, params).fetchall()
    finally:
        conn.close()

    BG = {'on_time': 'C6EFCE', 'late': 'FFC7CE',
          'absent': 'F2F2F2', 'early_leave': 'FFEB9C', 'present': 'DDEEFF'}

    for i, row in enumerate(rows, 3):
        s  = row['status']
        bg = BG.get(s)
        vals = [
            row['name_ar'], row['att_date'],
            row['check_in'] or '—', row['check_out'] or '—',
            row['late_min'] or 0, row['early_min'] or 0,
            STATUS_AR.get(s, s)
        ]
        for ci, v in enumerate(vals, 1):
            _cell(ws, i, ci, v, bg=bg)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ═══════════════════════════════════════════════════════════
#  EXCEL — PAYROLL EXPORT  (enhanced version)
# ═══════════════════════════════════════════════════════════

# -- Colors --
_XC = {
    'navy':    '1B2A4A',  # dark header
    'blue':    '1D6FAF',  # sub-header
    'lblue':   'D6E8F7',  # info background
    'green':   '197A3E',  # on-time text
    'lgreen':  'D1FAE5',  # on-time background
    'orange':  'B45309',  # late text
    'lorange': 'FEF3C7',  # late background
    'red':     'B91C1C',  # absent / deduction text
    'lred':    'FEE2E2',  # absent background
    'gray':    'F1F5F9',  # alternating row
    'dgray':   'CBD5E1',  # borders
    'gold':    'D97706',  # leave
    'lgold':   'FFFBEB',  # leave background
    'white':   'FFFFFF',
    'teal':    '0F766E',  # summary
    'lteal':   'CCFBF1',  # summary background
}

def _xfont(bold=False, size=10, color='000000', name='Arial'):
    return Font(bold=bold, size=size, color=color, name=name)

def _xfill(color):
    return PatternFill('solid', fgColor=color)

def _xalign(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, reading_order=2)

def _thin_border(color='CBD5E1'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _medium_border():
    s = Side(style='medium', color='1B2A4A')
    return Border(left=s, right=s, top=s, bottom=s)

def _xrow(ws, r, h):
    ws.row_dimensions[r].height = h

def _xcol(ws, cols_widths):
    for i, w in enumerate(cols_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _xc(ws, r, c, val, bold=False, size=10, fg='000000', bg=None,
        h='center', v='center', wrap=False, num_fmt=None, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = _xfont(bold=bold, size=size, color=fg)
    cell.alignment = _xalign(h=h, v=v, wrap=wrap)
    if bg:  cell.fill = _xfill(bg)
    if border: cell.border = _thin_border()
    if num_fmt: cell.number_format = num_fmt
    return cell

def _xmerge(ws, r, c1, c2, val, bold=False, size=10, fg='000000', bg=None,
             h='center', v='center', wrap=False, num_fmt=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = _xc(ws, r, c1, val, bold=bold, size=size, fg=fg, bg=bg,
             h=h, v=v, wrap=wrap, num_fmt=num_fmt, border=False)
    # draw medium border around merged block
    thin = Side(style='thin', color=_XC['dgray'])
    med  = Side(style='medium', color=_XC['navy'])
    c.border = Border(left=med, right=med, top=med, bottom=med)
    return c

# map violation causes → short label
_VTYPE_AR = {
    'late':       'Late Arrival',
    'early_leave':'Early Departure',
    'flex_hours': 'Missing Hours',
    'absent_1':   'Absent',
}

_OCC_WORDS = ['1st','2nd','3rd','4th','5th',
              '6th','7th','8th','9th','10th']
def _occ_ar(n):
    """Return ordinal string for occurrence number"""
    if 1 <= n <= len(_OCC_WORDS):
        return f'Occurrence {_OCC_WORDS[n-1]}'
    return f'Occurrence {n}'

_STATUS_AR = {
    'on_time':         ('On Time',       _XC['lgreen'],  _XC['green']),
    'late':            ('Late',          _XC['lorange'], _XC['orange']),
    'absent':          ('Absent',           _XC['lred'],    _XC['red']),
    'early_leave':     ('Early Departure',      _XC['lorange'], _XC['orange']),
    'leave_sick':      ('Sick Leave',    _XC['lgold'],   _XC['gold']),
    'leave_emergency': ('Emergency Leave',    _XC['lgold'],   _XC['gold']),
    'leave_annual':    ('Annual Leave',    _XC['lgold'],   _XC['gold']),
    'leave_official':  ('Official Holiday',    _XC['lgold'],   _XC['gold']),
}

def _status_info(s):
    return _STATUS_AR.get(s, (s, _XC['white'], '000000'))

def _ptype_ar(ptype, pvalue):
    if ptype == 'warning': return 'Warning'
    if ptype == 'percent': return f'{pvalue}% of daily'
    if ptype in ('day', 'warning_day'): return f'Deduct {pvalue} day(s)'
    return f'{ptype} {pvalue}'

def export_payroll_excel(year, month):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = get_db()
    try:
        emps   = conn.execute("SELECT * FROM employees ORDER BY name_ar").fetchall()
        prefix = f"{year}-{month:02d}-%"
        payroll_summary = []

        # ══════════════════════════════════════════════════════
        # sheet per employee
        # ══════════════════════════════════════════════════════
        for emp_row in emps:
            emp  = dict(emp_row)
            atts = conn.execute(
                "SELECT * FROM attendance WHERE employee_id=? AND att_date LIKE ? ORDER BY att_date",
                (emp['id'], prefix)).fetchall()
            vios = conn.execute(
                "SELECT * FROM violations WHERE employee_id=? AND vio_date LIKE ? ORDER BY vio_date",
                (emp['id'], prefix)).fetchall()

            total_ded = sum(v['deduction'] for v in vios)
            gross     = (emp['salary'] or 0) + (emp['housing'] or 0) + (emp['transport'] or 0) + (emp['commission'] or 0)
            gosi_ded  = _gosi(emp)
            net       = gross - total_ded - gosi_ded

            is_fixed = (emp['work_type'] == 'fixed')
            wtype_ar = 'Fixed Schedule' if is_fixed else 'Flexible Schedule'

            ws = wb.create_sheet(emp['name_en'][:28])
            ws.sheet_view.rightToLeft    = True
            ws.sheet_view.showGridLines  = False
            ws.print_area = 'A1:I50'

            # columns: Date | Check-in | Check-out | Late(min) | Early Exit(min) | Hours | Status | Violation | Deduction
            _xcol(ws, [13, 9, 9, 9, 12, 8, 13, 30, 12])

            r = 1
            # -- title bar --
            _xrow(ws, r, 46)
            _xmerge(ws, r, 1, 9,
                    f"Attendance Record  ▸  {emp['name_ar']}  ▸  {MONTHS_AR[month-1]} {year}",
                    bold=True, size=15, fg=_XC['white'], bg=_XC['navy'], h='center')
            r += 1

            # -- employee info bar --
            _xrow(ws, r, 22)
            _xmerge(ws, r, 1, 5,
                    f"Schedule Type: {wtype_ar}   |   Work Hours: {emp['work_start']} — {emp['work_end']}",
                    size=10, fg=_XC['navy'], bg=_XC['lblue'], h='right')
            _xmerge(ws, r, 6, 9,
                    f"Basic: {emp['salary']:,.0f}   Housing: {emp['housing']:,.0f}   "
                    f"Transport: {emp['transport']:,.0f}   Commission: {emp['commission']:,.0f}   (SR)",
                    size=10, fg=_XC['navy'], bg=_XC['lblue'], h='right')
            r += 1

            # -- column headers --────────────────────────────────
            _xrow(ws, r, 26)
            hdrs = ['Date', 'Check-in', 'Check-out', 'Late\n(min)', 'Early Exit\n(min)',
                    'Hours', 'Status', 'Violation', 'Deduction\n(SR)']
            for ci, h in enumerate(hdrs, 1):
                _xc(ws, r, ci, h, bold=True, size=10, fg=_XC['white'],
                    bg=_XC['blue'], wrap=True)
            r += 1

            # -- attendance rows --
            occ_count = {}   # {vtype: occurrence count this month}
            for att in atts:
                att   = dict(att)
                s     = att['status']
                label, row_bg, txt_color = _status_info(s)
                day_vios  = [v for v in vios if v['vio_date'] == att['att_date']]
                day_ded   = sum(v['deduction'] for v in day_vios)

                vio_parts = []
                for v in day_vios:
                    vt = v['vtype']
                    occ_count[vt] = occ_count.get(vt, 0) + 1
                    occ_label = _occ_ar(occ_count[vt])
                    penalty   = _ptype_ar(v['ptype'], v['pvalue'])
                    vio_parts.append(
                        f"{_VTYPE_AR.get(vt, vt)}  ◂  {occ_label}  ◂  {penalty}"
                    )
                vio_text  = '\n'.join(vio_parts) if vio_parts else '—'

                late_m  = att['late_min']  or 0
                early_m = att['early_min'] or 0
                hours   = round(att['total_hours'] or 0, 2)

                _xrow(ws, r, 18 if not vio_parts else 20)
                row_vals = [
                    att['att_date'],
                    att['check_in']  or '—',
                    att['check_out'] or '—',
                    late_m  if late_m  > 0 else '—',
                    early_m if early_m > 0 else '—',
                    hours   if hours   > 0 else '—',
                    label,
                    vio_text,
                    round(day_ded, 2) if day_ded else '—',
                ]
                for ci, val in enumerate(row_vals, 1):
                    is_vio = (ci == 8)
                    is_ded = (ci == 9)
                    cell_bg = row_bg
                    cell_fg = txt_color if ci == 7 else ('000000' if not is_ded else (_XC['red'] if day_ded > 0 else '000000'))
                    c = _xc(ws, r, ci, val, fg=cell_fg, bg=cell_bg,
                             size=9 if is_vio else 10,
                             wrap=is_vio, h='right' if is_vio else 'center',
                             num_fmt='#,##0.00' if is_ded and isinstance(val, float) else None)
                r += 1

            # -- spacer --
            r += 1

            # -- monthly summary (4 cols x 2 rows + net) --
            _xrow(ws, r, 26)
            _xmerge(ws, r, 1, 9, 'Monthly Summary',
                    bold=True, size=12, fg=_XC['white'], bg=_XC['teal'], h='center')
            r += 1

            present_cnt = sum(1 for a in atts if not a['status'].startswith('leave_') and a['status'] != 'absent')
            absent_cnt  = sum(1 for a in atts if a['status'] == 'absent')
            late_cnt    = sum(1 for a in atts if a['status'] == 'late')
            early_cnt   = sum(1 for a in atts if a['status'] == 'early_leave')
            total_late  = sum(a['late_min']  or 0 for a in atts)
            total_early = sum(a['early_min'] or 0 for a in atts)

            # two rows of stats
            stat_rows = [
                [('Attendance Days', present_cnt, _XC['lgreen']),
                  ('Absent Days',     absent_cnt,  _XC['lred']),
                  ('Late Days',      late_cnt,    _XC['lorange']),
                  ('Early Exits',    early_cnt,   _XC['lorange'])],
                [('Total Late (min)',    total_late,  _XC['lorange']),
                  ('Early Exit (min)',   total_early, _XC['lorange']),
                  ('Total Deductions',   round(total_ded, 2), _XC['lred']),
                  ('GOSI Insurance',     round(gosi_ded, 2), _XC['gray'])],
            ]
            col_map = [(1,2), (3,4), (5,6), (7,9)]
            for stat_row in stat_rows:
                _xrow(ws, r, 22)
                for (c1, c2), (lbl, val, bg) in zip(col_map, stat_row):
                    txt = f"{lbl}\n{val:,}" if isinstance(val, int) else f"{lbl}\n{val:,.2f} SR"
                    _xmerge(ws, r, c1, c2, txt, size=9, bg=bg, h='center', wrap=True)
                r += 1

            # gross salary row
            _xrow(ws, r, 22)
            _xmerge(ws, r, 1, 4, f"Gross Salary:  {gross:,.2f} SR",
                    bold=True, size=11, fg=_XC['navy'], bg=_XC['lblue'], h='right')
            _xmerge(ws, r, 5, 9, f"— Deductions {total_ded:,.2f}  — Insurance {gosi_ded:,.2f}",
                    size=10, fg=_XC['red'], bg=_XC['lred'], h='center')
            r += 1

            # net salary row
            _xrow(ws, r, 30)
            net_bg = _XC['lgreen'] if net >= gross * 0.85 else (_XC['lorange'] if net >= gross * 0.7 else _XC['lred'])
            net_fg = _XC['green']  if net >= gross * 0.85 else (_XC['orange']  if net >= gross * 0.7 else _XC['red'])
            _xmerge(ws, r, 1, 9, f"Net Salary:   {net:,.2f} SR",
                    bold=True, size=14, fg=net_fg, bg=net_bg, h='center')
            r += 1

            # signature area
            r += 2
            _xrow(ws, r, 20)
            for c1, c2, lbl in [(1, 3, 'Employee Signature'), (4, 6, 'Supervisor Signature'), (7, 9, 'Management Approval')]:
                _xmerge(ws, r, c1, c2, lbl, bold=True, size=10, fg=_XC['navy'], bg=_XC['gray'], h='center')
            r += 2
            _xrow(ws, r, 20)
            for c1, c2 in [(1, 3), (4, 6), (7, 9)]:
                _xmerge(ws, r, c1, c2, '─' * 28, size=9, fg=_XC['dgray'], h='center')

            payroll_summary.append({
                **emp,
                'total_ded': total_ded,
                'gosi_ded':  gosi_ded,
                'other_ded': gosi_ded,
                'gross':     gross,
                'net':       net,
                'days':      len(atts),
                'present':   present_cnt,
                'late_days': late_cnt,
                'absent':    absent_cnt,
            })

        # ══════════════════════════════════════════════════════
        # main payroll sheet (first sheet)
        # ══════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Payroll', 0)
        ws2.sheet_view.rightToLeft   = True
        ws2.sheet_view.showGridLines = False

        # columns: # | Employee | Schedule | Basic | Housing | Transport | Commission | Gross | Deductions | Insurance | Net | Notes
        _xcol(ws2, [5, 22, 11, 12, 11, 11, 10, 13, 11, 12, 14, 14])
        N = 12  # number of columns

        # -- main title --
        _xrow(ws2, 1, 52)
        _xmerge(ws2, 1, 1, N,
                f"Payroll  ▸  {MONTHS_AR[month-1]} {year}",
                bold=True, size=20, fg=_XC['white'], bg=_XC['navy'], h='center')

        # -- info bar --
        _xrow(ws2, 2, 24)
        _xmerge(ws2, 2, 1, 6,
                f"Prepared: {date.today().strftime('%Y/%m/%d')}",
                size=10, fg=_XC['navy'], bg=_XC['lblue'], h='right')
        _xmerge(ws2, 2, 7, N,
                f"Total Employees: {len(payroll_summary)}",
                size=10, fg=_XC['navy'], bg=_XC['lblue'], h='right')

        # -- column headers --
        _xrow(ws2, 3, 30)
        pay_hdrs = ['#', 'Employee Name', 'Schedule Type',
                    'Basic Salary', 'Housing', 'Transport', 'Commission',
                    'Gross', 'Violation Deductions', 'GOSI (10.75%)', 'Net Salary', 'Notes']
        for ci, h in enumerate(pay_hdrs, 1):
            _xc(ws2, 3, ci, h, bold=True, size=10, fg=_XC['white'], bg=_XC['blue'], wrap=True)

        # -- employee rows --
        tgross = tnet = tded = tgosi = 0.0
        for idx, pd in enumerate(payroll_summary, 1):
            row_r  = idx + 3
            alt_bg = _XC['gray'] if idx % 2 == 0 else _XC['white']
            net_bg = _XC['lgreen'] if pd['net'] >= pd['gross'] * 0.85 else (
                     _XC['lorange'] if pd['net'] >= pd['gross'] * 0.7 else _XC['lred'])

            wtype_ar = 'Fixed' if pd['work_type'] == 'fixed' else 'Flex'
            _xrow(ws2, row_r, 22)

            vals = [
                (idx,               alt_bg, 'center'),
                (pd['name_ar'],     alt_bg, 'right'),
                (wtype_ar,          alt_bg, 'center'),
                (pd['salary'],      alt_bg, 'center'),
                (pd['housing'],     alt_bg, 'center'),
                (pd['transport'],   alt_bg, 'center'),
                (pd['commission'],  alt_bg, 'center'),
                (pd['gross'],       alt_bg, 'center'),
                (pd['total_ded'],   _XC['lred']   if pd['total_ded'] > 0 else alt_bg, 'center'),
                (pd['gosi_ded'],    alt_bg, 'center'),
                (pd['net'],         net_bg, 'center'),
                ('',                alt_bg, 'center'),
            ]
            for ci, (val, bg, align) in enumerate(vals, 1):
                num_f = '#,##0.00' if ci in (4,5,6,7,8,9,10,11) else None
                _xc(ws2, row_r, ci, val, bg=bg, h=align, size=10, num_fmt=num_f)

            tgross += pd['gross']
            tnet   += pd['net']
            tded   += pd['total_ded']
            tgosi  += pd['gosi_ded']

        # -- totals row --
        tot = len(payroll_summary) + 4
        _xrow(ws2, tot, 30)
        _xmerge(ws2, tot, 1, 3, 'Grand Total',
                bold=True, size=12, fg=_XC['white'], bg=_XC['navy'], h='center')
        for ci in range(4, N + 1):
            _xc(ws2, tot, ci, '', bold=True, bg=_XC['navy'])
        for ci, val in [(8, tgross), (9, tded), (10, tgosi), (11, tnet)]:
            _xc(ws2, tot, ci, val, bold=True, size=11, fg=_XC['white'],
                bg=_XC['navy'], num_fmt='#,##0.00')

        # -- savings row (net vs gross) --
        _xrow(ws2, tot + 1, 16)
        pct = round((1 - tded / tgross) * 100, 1) if tgross else 100
        _xmerge(ws2, tot + 1, 1, N,
                f"Deduction rate from gross:  {pct}%   |   Net payroll:  {tnet:,.2f} SR",
                size=10, fg=_XC['navy'], bg=_XC['lteal'], h='center')

        # -- signature area --
        sig = tot + 4
        _xrow(ws2, sig, 22)
        for c1, c2, lbl in [(1, 4, 'Prepared by:   _________________________'),
                             (5, 8, 'Reviewed by:   _________________________'),
                             (9, N, 'Approved by:   _________________________')]:
            _xmerge(ws2, sig, c1, c2, lbl, bold=True, size=11, fg=_XC['navy'], h='center')

    finally:
        conn.close()

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_gosi_excel(year, month):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'GOSI'
    ws.sheet_view.rightToLeft   = True
    ws.sheet_view.showGridLines = False
    month_ar = MONTHS_AR[month - 1]

    _xcol(ws, [6, 30, 20, 18, 18, 18, 18])

    # title
    _xrow(ws, 1, 44)
    _xmerge(ws, 1, 1, 7,
            f"Social Insurance Report (GOSI)  —  {month_ar} {year}",
            bold=True, size=14, fg=_XC['white'], bg=_XC['navy'], h='center')

    # column headers
    _xrow(ws, 2, 26)
    for ci, h in enumerate(['#', 'Employee Name', 'Employee Code',
                             'Basic Salary', 'Housing', 'Transport',
                             'GOSI (10.75%)'], 1):
        _xc(ws, 2, ci, h, bold=True, size=10,
            fg=_XC['white'], bg=_XC['blue'])

    conn = get_db()
    try:
        emps = conn.execute("SELECT * FROM employees ORDER BY name_ar").fetchall()
        total_gosi = 0.0
        for i, emp_row in enumerate(emps, 1):
            emp      = dict(emp_row)
            gosi_ded = _gosi(emp)
            total_gosi += gosi_ded
            bg = _XC['lrow1'] if i % 2 else _XC['white']
            _xrow(ws, i + 2, 22)
            vals = [i, emp['name_ar'], emp['emp_code'] or '—',
                    emp['salary'], emp['housing'], emp['transport'],
                    round(gosi_ded, 2)]
            for ci, v in enumerate(vals, 1):
                _xc(ws, i + 2, ci, v, size=10, bg=bg,
                    fg=_XC['red'] if ci == 7 and gosi_ded == 0 else _XC['dark'])

        # total row
        r = len(emps) + 3
        _xrow(ws, r, 26)
        _xmerge(ws, r, 1, 6, 'Total',
                bold=True, size=11, fg=_XC['white'], bg=_XC['navy'], h='right')
        _xc(ws, r, 7, round(total_gosi, 2),
            bold=True, size=11, fg=_XC['white'], bg=_XC['navy'])
    finally:
        conn.close()

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ═══════════════════════════════════════════════════════════
#  FLASK ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')

# ── Dashboard ────────────────────────────────────────────
@app.route('/api/stats/today')
def api_stats_today():
    today = str(date.today())
    conn = get_db()
    try:
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM employees").fetchone()['c']
        row = conn.execute("""
            SELECT
                SUM(CASE WHEN status='on_time'    THEN 1 ELSE 0 END) AS on_time,
                SUM(CASE WHEN status='late'        THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN status='absent'      THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN status='early_leave' THEN 1 ELSE 0 END) AS early_leave,
                SUM(CASE WHEN status='present'     THEN 1 ELSE 0 END) AS present
            FROM attendance WHERE att_date=?""", (today,)).fetchone()
    finally:
        conn.close()
    return jsonify({
        'date': today, 'total_employees': total,
        'on_time':     row['on_time']     or 0,
        'late':        row['late']        or 0,
        'absent':      row['absent']      or 0,
        'early_leave': row['early_leave'] or 0,
        'present':     row['present']     or 0,
    })

@app.route('/api/stats/today/detail')
def api_stats_today_detail():
    status = request.args.get('status', '')
    today  = str(date.today())
    conn   = get_db()
    try:
        if status == 'absent':
            # absent = employees with no record today or status=absent
            rows = conn.execute("""
                SELECT e.name_ar, e.emp_code, a.check_in, a.check_out, a.late_min, a.status
                FROM employees e
                LEFT JOIN attendance a ON a.employee_id=e.id AND a.att_date=?
                WHERE a.status='absent' OR a.id IS NULL
                ORDER BY e.name_ar
            """, (today,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT e.name_ar, e.emp_code, a.check_in, a.check_out, a.late_min, a.status
                FROM attendance a
                JOIN employees e ON e.id=a.employee_id
                WHERE a.att_date=? AND a.status=?
                ORDER BY e.name_ar
            """, (today, status)).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/stats/month/detail')
def api_stats_month_detail():
    y      = request.args.get('year',  date.today().year,  type=int)
    m      = request.args.get('month', date.today().month, type=int)
    dtype  = request.args.get('type', '')
    prefix = f"{y}-{m:02d}-%"
    conn   = get_db()
    try:
        if dtype == 'violations':
            rows = conn.execute("""
                SELECT e.name_ar, e.emp_code, COUNT(*) AS cnt,
                       SUM(v.deduction) AS total_ded
                FROM violations v JOIN employees e ON e.id=v.employee_id
                WHERE v.vio_date LIKE ?
                GROUP BY v.employee_id ORDER BY cnt DESC
            """, (prefix,)).fetchall()
        elif dtype == 'deductions':
            rows = conn.execute("""
                SELECT e.name_ar, e.emp_code,
                       COALESCE(SUM(v.deduction),0) AS total_ded,
                       COUNT(v.id) AS cnt
                FROM employees e
                LEFT JOIN violations v ON v.employee_id=e.id AND v.vio_date LIKE ?
                GROUP BY e.id
                HAVING total_ded > 0
                ORDER BY total_ded DESC
            """, (prefix,)).fetchall()
        elif dtype == 'absent':
            rows = conn.execute("""
                SELECT e.name_ar, e.emp_code, COUNT(*) AS cnt
                FROM attendance a JOIN employees e ON e.id=a.employee_id
                WHERE a.att_date LIKE ? AND a.status='absent'
                GROUP BY a.employee_id ORDER BY cnt DESC
            """, (prefix,)).fetchall()
        else:
            rows = []
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/stats/month')
def api_stats_month():
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        att = conn.execute("""
            SELECT COUNT(*) AS total,
                SUM(CASE WHEN status='late'        THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN status='absent'      THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN status='early_leave' THEN 1 ELSE 0 END) AS early_leave,
                SUM(COALESCE(late_min,0)) AS late_min_total
            FROM attendance WHERE att_date LIKE ?""", (prefix,)).fetchone()
        vio = conn.execute("""
            SELECT COALESCE(SUM(deduction),0) AS d, COUNT(*) AS c
            FROM violations WHERE vio_date LIKE ?""", (prefix,)).fetchone()
    finally:
        conn.close()
    return jsonify({
        'year': y, 'month': m,
        'total':          att['total']         or 0,
        'late':           att['late']           or 0,
        'absent':         att['absent']         or 0,
        'early_leave':    att['early_leave']    or 0,
        'late_min_total': att['late_min_total'] or 0,
        'violations':     vio['c']              or 0,
        'total_deductions': float(vio['d']      or 0),
    })

@app.route('/api/stats/trend')
@login_required
def api_stats_trend():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT att_date,
                SUM(CASE WHEN status='on_time'  THEN 1 ELSE 0 END) as on_time,
                SUM(CASE WHEN status='late'      THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status='absent'    THEN 1 ELSE 0 END) as absent
            FROM attendance
            WHERE att_date >= date('now', '-30 days')
            GROUP BY att_date ORDER BY att_date
        """).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/notifications')
@hr_required
def api_notifications():
    conn = get_db()
    try:
        excuses     = conn.execute("SELECT COUNT(*) FROM excuse_requests    WHERE status='pending'").fetchone()[0]
        leaves      = conn.execute("SELECT COUNT(*) FROM leaves             WHERE status='pending'").fetchone()[0]
        overtime    = conn.execute("SELECT COUNT(*) FROM overtime_requests  WHERE status='pending'").fetchone()[0]
        att_reqs    = conn.execute("SELECT COUNT(*) FROM attendance_requests WHERE status='pending'").fetchone()[0]
    finally:
        conn.close()
    return jsonify({'excuses': excuses, 'leaves': leaves, 'overtime': overtime,
                    'att_requests': att_reqs, 'total': excuses+leaves+overtime+att_reqs})

@app.route('/api/settings', methods=['GET'])
@login_required
def api_settings_get():
    conn = get_db()
    try:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    finally:
        conn.close()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['PUT'])
@hr_required
def api_settings_put():
    data = request.get_json() or {}
    conn = get_db()
    try:
        for k, v in data.items():
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, str(v)))
        conn.commit()
    finally:
        conn.close()
    # invalidate TTLock token cache when credentials change
    tt_keys = {'tt_client_id','tt_client_secret','tt_username','tt_password'}
    if any(k in tt_keys for k in data):
        global _tt_cache
        _tt_cache = {'token': None, 'exp': 0}
    audit_log('edit_settings', 'settings', str(list(data.keys())))
    return jsonify({'ok': True})

@app.route('/api/ttlock/test', methods=['POST'])
@hr_required
def api_ttlock_test():
    """Test TTLock connection — returns token status and lock count."""
    _tt_cache['token'] = None  # force fresh token
    cid, _, usr, _ = _tt_creds()
    if not cid or not usr:
        return jsonify({'ok': False, 'msg': 'Credentials not configured — enter Client ID, Username, and Password first.'})
    token, err = tt_get_token(return_error=True)
    if not token:
        detail = f' — {err}' if err and err != 'missing_credentials' else ''
        return jsonify({'ok': False, 'msg': f'Authentication failed{detail}', 'server': TTBASE})
    locks = tt_get_locks(token)
    return jsonify({'ok': True, 'msg': f'Connected successfully! Found {len(locks)} lock(s).', 'locks': len(locks)})

@app.route('/api/system/info')
@hr_required
def api_system_info():
    import os
    db_path = DB_PATH
    db_exists = os.path.exists(db_path)
    db_size = os.path.getsize(db_path) if db_exists else 0
    cid, _, usr, _ = _tt_creds()
    return jsonify({
        'db_path': db_path,
        'db_exists': db_exists,
        'db_size_kb': round(db_size / 1024, 1),
        'ttlock_cid_set': bool(cid),
        'ttlock_usr': usr,
        'ttlock_server': TTBASE,
    })

@app.route('/api/ttlock/debug-records')
@hr_required
def api_ttlock_debug():
    """Show raw TTLock usernames for today — helps diagnose name mismatch."""
    token, err = tt_get_token(return_error=True)
    if not token:
        return jsonify({'ok': False, 'msg': f'Auth failed: {err}'})
    target = date.today()
    TZ_OFFSET = 3 * 3600
    start_ms = int(datetime(target.year, target.month, target.day, 0, 0, 0).timestamp() * 1000) - TZ_OFFSET * 1000
    end_ms   = int(datetime(target.year, target.month, target.day, 23, 59, 59).timestamp() * 1000) - TZ_OFFSET * 1000
    locks = tt_get_locks(token)
    lock_ids = [l.get('lockId') for l in locks]
    raw_responses = []
    by_user = {}
    ts = int(datetime.now().timestamp() * 1000)
    cid, _, _, _ = _tt_creds()
    for lid in lock_ids:
        if not lid: continue
        try:
            r = requests.get(f"{TTBASE}/v3/lockRecord/list", params={
                'clientId': cid, 'accessToken': token,
                'lockId': lid, 'startDate': start_ms,
                'endDate': end_ms, 'pageNo': 1,
                'pageSize': 20, 'date': ts
            }, timeout=15).json()
            raw_responses.append({'lockId': lid, 'response': r})
            for rec in r.get('list', []):
                uname = (rec.get('username') or '').strip()
                ts_ms = rec.get('successDate', 0)
                if uname and ts_ms:
                    by_user.setdefault(uname, []).append(
                        datetime.fromtimestamp(ts_ms / 1000).strftime('%H:%M:%S'))
        except Exception as e:
            raw_responses.append({'lockId': lid, 'error': str(e)})
    return jsonify({
        'date': str(target),
        'start_ms': start_ms, 'end_ms': end_ms,
        'locks_found': lock_ids,
        'users_found': list(by_user.keys()),
        'records': by_user,
        'raw_api_responses': raw_responses
    })

@app.route('/api/attendance/recent')
def api_attendance_recent():
    limit = request.args.get('limit', 15, type=int)
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT a.*, e.name_ar
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            ORDER BY a.att_date DESC, a.check_in DESC
            LIMIT ?""", (limit,)).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/clear-fake-data', methods=['POST'])
@hr_required
def api_clear_fake_data():
    """Clear all attendance records and violations generated without TTLock connection"""
    conn = get_db()
    try:
        a = conn.execute("DELETE FROM attendance").rowcount
        v = conn.execute("DELETE FROM violations").rowcount
        conn.execute("DELETE FROM vio_counts")
        conn.commit()
        return jsonify({'ok': True, 'msg': f'Cleared {a} attendance records and {v} violations'})
    finally:
        conn.close()

@app.route('/api/run', methods=['POST'])
def api_run():
    data = request.get_json(silent=True) or {}
    d_str = data.get('date')
    try:
        target = datetime.strptime(d_str, '%Y-%m-%d').date() if d_str else None
        process_day(target)
        audit_log('manual_run', 'attendance', d_str or str(date.today()))
        return jsonify({'ok': True, 'msg': 'Attendance processed successfully'})
    except Exception as e:
        logger.error(f"Manual run error: {e}", exc_info=True)
        return jsonify({'ok': False, 'msg': str(e)}), 500

# ── Attendance ────────────────────────────────────────────
@app.route('/api/attendance')
def api_attendance():
    y   = request.args.get('year',  date.today().year,  type=int)
    m   = request.args.get('month', date.today().month, type=int)
    eid = request.args.get('emp_id', type=int)
    prefix = f"{y}-{m:02d}-%"
    q = """SELECT a.*, e.name_ar
           FROM attendance a JOIN employees e ON e.id=a.employee_id
           WHERE a.att_date LIKE ?"""
    params = [prefix]
    if eid:
        q += " AND a.employee_id=?"
        params.append(eid)
    q += " ORDER BY a.att_date DESC, e.name_ar"
    conn = get_db()
    try:
        rows = conn.execute(q, params).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/attendance/export')
def api_attendance_export():
    y   = request.args.get('year',  date.today().year,  type=int)
    m   = request.args.get('month', date.today().month, type=int)
    eid = request.args.get('emp_id', type=int)
    out = export_attendance_excel(y, m, eid)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"attendance_{y}_{m:02d}.xlsx")

# ── Employees ─────────────────────────────────────────────
@app.route('/api/employees', methods=['GET'])
def api_emps_get():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM employees ORDER BY name_ar").fetchall()
        year = date.today().year
        result = []
        for r in rows:
            d = dict(r)
            d['leave_balance'] = _leave_balance(conn, r['id'], year)
            yrs = _years_of_service(d.get('hire_date'))
            d['years_of_service'] = yrs
            d['total_salary'] = (d.get('salary') or 0) + (d.get('housing') or 0) + (d.get('transport') or 0)
            d['eosb'] = _eosb(d.get('salary'), yrs)
            d['alert'] = _iqama_alert(d.get('iqama_expiry'))
            result.append(d)
    finally:
        conn.close()
    return jsonify(result)

@app.route('/api/employees/<int:eid>/leave-balance')
@login_required
def api_leave_balance(eid):
    year = request.args.get('year', date.today().year, type=int)
    conn = get_db()
    try:
        bal = _leave_balance(conn, eid, year)
    finally:
        conn.close()
    return jsonify(bal)

@app.route('/api/employees', methods=['POST'])
def api_emps_post():
    d = request.get_json(silent=True) or {}
    if not d.get('name_ar') or not d.get('name_en'):
        return jsonify({'error': 'name_ar and name_en are required'}), 400
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO employees
                (name_ar,name_en,email,salary,housing,transport,commission,
                 other_ded,work_type,work_start,work_end,weekly_hours,annual_leave_days,emp_code,weekend_days,
                 nationality,department,job_title,status,hire_date,national_id,phone,
                 direct_manager,notes,contract_type,contract_start,contract_end,
                 probation_end,iqama_expiry,gosi_number,iban,medical_insurance)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d['name_ar'], d['name_en'], d.get('email'),
             d.get('salary',0) or 0, d.get('housing',0) or 0, d.get('transport',0) or 0,
             d.get('commission') or 0, d.get('other_ded',0) or 0,
             d.get('work_type','fixed'), d.get('work_start','08:00'),
             d.get('work_end','17:00'), d.get('weekly_hours',40),
             d.get('annual_leave_days', 21), d.get('emp_code') or None,
             d.get('weekend_days', '5,6'),
             d.get('nationality') or None, d.get('department') or None,
             d.get('job_title') or None, d.get('status','active'),
             d.get('hire_date') or None, d.get('national_id') or None,
             d.get('phone') or None, d.get('direct_manager') or None,
             d.get('notes') or None, d.get('contract_type','permanent'),
             d.get('contract_start') or None, d.get('contract_end') or None,
             d.get('probation_end') or None, d.get('iqama_expiry') or None,
             d.get('gosi_number') or None, d.get('iban') or None,
             d.get('medical_insurance') or None))
        conn.commit()
        audit_log('create_employee', 'employee', d['name_ar'])
        return jsonify({'ok': True, 'msg': 'Employee added successfully'})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'English name already in use'}), 400
    finally:
        conn.close()

@app.route('/api/employees/<int:eid>', methods=['GET'])
def api_emp_get(eid):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    finally:
        conn.close()
    if not row:
        return jsonify({'error': 'Employee not found'}), 404
    return jsonify(dict(row))

@app.route('/api/employees/<int:eid>', methods=['PUT'])
def api_emp_put(eid):
    d = request.get_json(silent=True) or {}
    allowed = ['name_ar','name_en','email','salary','housing','transport',
               'commission','other_ded','work_type','work_start','work_end',
               'weekly_hours','annual_leave_days','emp_code','weekend_days',
               'nationality','department','job_title','status','hire_date',
               'national_id','phone','direct_manager','notes','contract_type',
               'contract_start','contract_end','probation_end','iqama_expiry',
               'gosi_number','iban','medical_insurance']
    updates = {k: d[k] for k in allowed if k in d}
    if not updates:
        return jsonify({'error': 'No fields to update'}), 400
    schedule_fields = {'work_type','work_start','work_end','weekly_hours','weekend_days'}
    effective_date  = d.get('schedule_effective_date', '')
    sql = f"UPDATE employees SET {', '.join(k+'=?' for k in updates)} WHERE id=?"
    conn = get_db()
    try:
        conn.execute(sql, list(updates.values()) + [eid])
        # if schedule changed and effective date specified — log it in history
        if effective_date and any(k in schedule_fields for k in updates):
            emp = conn.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
            if emp:
                conn.execute("""
                    INSERT INTO schedule_history
                        (employee_id, effective_date, work_type, work_start, work_end, weekly_hours, weekend_days)
                    VALUES (?,?,?,?,?,?,?)
                """, (eid, effective_date,
                      updates.get('work_type',    emp['work_type']),
                      updates.get('work_start',   emp['work_start']),
                      updates.get('work_end',     emp['work_end']),
                      updates.get('weekly_hours', emp['weekly_hours']),
                      updates.get('weekend_days', emp.get('weekend_days','5,6'))))
        conn.commit()
        audit_log('edit_employee', 'employee', str(eid), details=str(list(updates.keys())))
    finally:
        conn.close()
    return jsonify({'ok': True, 'msg': 'Employee data updated'})

@app.route('/api/employees/<int:eid>', methods=['DELETE'])
def api_emp_delete(eid):
    conn = get_db()
    try:
        conn.execute("DELETE FROM employees WHERE id=?", (eid,))
        conn.commit()
        audit_log('delete_employee', 'employee', str(eid))
    finally:
        conn.close()
    return jsonify({'ok': True, 'msg': 'Employee deleted'})

# ── Payroll ───────────────────────────────────────────────
@app.route('/api/payroll')
def api_payroll():
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        emps = conn.execute(
            "SELECT * FROM employees ORDER BY name_ar").fetchall()
        result = []
        for emp_row in emps:
            emp = dict(emp_row)
            vd = conn.execute(
                "SELECT COALESCE(SUM(deduction),0) AS d FROM violations "
                "WHERE employee_id=? AND vio_date LIKE ?",
                (emp['id'], prefix)).fetchone()
            ad = conn.execute("""
                SELECT COUNT(*) AS total,
                    SUM(CASE WHEN status='absent' THEN 1 ELSE 0 END) AS absent,
                    SUM(CASE WHEN status='late'   THEN 1 ELSE 0 END) AS late
                FROM attendance WHERE employee_id=? AND att_date LIKE ?""",
                (emp['id'], prefix)).fetchone()
            gross    = (emp['salary'] or 0) + (emp['housing'] or 0) + (emp['transport'] or 0) + (emp['commission'] or 0)
            gosi_ded = _gosi(emp)
            net      = gross - (vd['d'] or 0) - gosi_ded
            result.append({
                **emp,
                'total_ded': round(vd['d'] or 0, 2),
                'gosi_ded':  gosi_ded,
                'other_ded': gosi_ded,
                'gross':     round(gross, 2),
                'net':       round(net, 2),
                'days':      ad['total']  or 0,
                'absent':    ad['absent'] or 0,
                'late':      ad['late']   or 0,
            })
    finally:
        conn.close()
    return jsonify(result)

@app.route('/api/payroll/export')
def api_payroll_export():
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    out = export_payroll_excel(y, m)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"payroll_{y}_{m:02d}.xlsx")

@app.route('/api/gosi/export')
@hr_required
def api_gosi_export():
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    out = export_gosi_excel(y, m)
    audit_log('export_gosi', details=f"{y}-{m:02d}")
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"gosi_{y}_{m:02d}.xlsx")

@app.route('/api/audit-logs')
@hr_required
def api_audit_logs():
    if not SUPABASE_KEY:
        return jsonify([])
    limit  = request.args.get('limit', 100, type=int)
    offset = request.args.get('offset', 0, type=int)
    try:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/audit_logs",
            headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'},
            params={'order': 'created_at.desc', 'limit': limit, 'offset': offset},
            timeout=8)
        return jsonify(resp.json())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/violations')
def api_violations():
    y   = request.args.get('year',  date.today().year,  type=int)
    m   = request.args.get('month', date.today().month, type=int)
    eid = request.args.get('emp_id', type=int)
    q = """SELECT v.*, e.name_ar
           FROM violations v JOIN employees e ON e.id=v.employee_id
           WHERE v.vio_date LIKE ?"""
    params = [f"{y}-{m:02d}-%"]
    if eid:
        q += " AND v.employee_id=?"
        params.append(eid)
    conn = get_db()
    try:
        rows = conn.execute(q + " ORDER BY v.vio_date DESC", params).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

# ═══════════════════════════════════════════════════════════
#  AUTH ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    d  = request.get_json(silent=True) or {}
    un = (d.get('username') or '').strip()
    pw = d.get('password', '')
    if not un or not pw:
        return jsonify({'error': 'Username and password are required'}), 400
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM users WHERE username=? AND password_hash=?",
            (un, _hash(pw))).fetchone()
        if not row:
            audit_log('login_failed', target_name=un, status='failed')
            return jsonify({'error': 'Invalid login credentials'}), 401
        session['user_id']   = row['id']
        session['username']  = row['username']
        session['role']      = row['role']
        session['employee_id'] = row['employee_id']
        audit_log('login', details=f"role={row['role']}")
        return jsonify({'ok': True, 'role': row['role'], 'username': row['username'],
                        'employee_id': row['employee_id']})
    finally:
        conn.close()

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    audit_log('logout')
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/me')
def api_me():
    if 'user_id' not in session:
        return jsonify({'logged_in': False})
    emp_id = session.get('employee_id')
    emp_data = None
    if emp_id:
        conn = get_db()
        try:
            row = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
            if row:
                emp_data = dict(row)
        finally:
            conn.close()
    return jsonify({'logged_in': True, 'role': session.get('role'),
                    'username': session.get('username'),
                    'employee_id': emp_id,
                    'employee': emp_data})

@app.route('/api/users', methods=['GET'])
@hr_required
def api_users_get():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT u.id, u.username, u.role, u.created_at,
                   e.name_ar, e.name_en
            FROM users u LEFT JOIN employees e ON e.id=u.employee_id
            ORDER BY u.role, u.username
        """).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/users', methods=['POST'])
@hr_required
def api_users_post():
    d = request.get_json(silent=True) or {}
    un = (d.get('username') or '').strip()
    pw = d.get('password', '').strip()
    role = d.get('role', 'employee')
    emp_id = d.get('employee_id') or None
    if not un or not pw:
        return jsonify({'error': 'Username and password are required'}), 400
    if role not in ('hr', 'manager', 'employee'):
        return jsonify({'error': 'Invalid role'}), 400
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, role, employee_id) VALUES (?,?,?,?)",
            (un, _hash(pw), role, emp_id))
        conn.commit()
        audit_log('create_user', 'user', un, details=f"role={role}")
        return jsonify({'ok': True, 'msg': 'User created'})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already in use'}), 400
    finally:
        conn.close()

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@hr_required
def api_user_delete(uid):
    if uid == session.get('user_id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    conn = get_db()
    try:
        conn.execute("DELETE FROM users WHERE id=?", (uid,))
        conn.commit()
        audit_log('delete_user', 'user', str(uid))
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/users/<int:uid>/password', methods=['PUT'])
@hr_required
def api_user_password(uid):
    d = request.get_json(silent=True) or {}
    pw = d.get('password', '').strip()
    if not pw or len(pw) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400
    conn = get_db()
    try:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?", (_hash(pw), uid))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  EXCUSE ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/api/excuses', methods=['GET'])
@login_required
def api_excuses_get():
    y   = request.args.get('year',  date.today().year,  type=int)
    m   = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        role   = session.get('role')
        emp_id = session.get('employee_id')
        if role in ('hr', 'manager'):
            rows = conn.execute("""
                SELECT ex.*, e.name_ar, e.name_en
                FROM excuse_requests ex
                JOIN employees e ON e.id=ex.employee_id
                WHERE ex.att_date LIKE ?
                ORDER BY ex.submitted_at DESC
            """, (prefix,)).fetchall()
        else:
            if not emp_id:
                return jsonify([])
            rows = conn.execute("""
                SELECT ex.*, e.name_ar, e.name_en
                FROM excuse_requests ex
                JOIN employees e ON e.id=ex.employee_id
                WHERE ex.employee_id=? AND ex.att_date LIKE ?
                ORDER BY ex.submitted_at DESC
            """, (emp_id, prefix)).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/excuses', methods=['POST'])
@login_required
def api_excuses_post():
    d       = request.get_json(silent=True) or {}
    emp_id  = session.get('employee_id')
    role    = session.get('role')
    # HR can submit for any employee
    if role in ('hr', 'manager'):
        emp_id = d.get('employee_id', emp_id)
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    att_date   = d.get('att_date', '')
    vtype      = d.get('vtype', 'late')
    reason     = (d.get('reason') or '').strip()
    attachment = d.get('attachment', '')
    att_name   = d.get('attachment_name', '')
    if not att_date or not reason:
        return jsonify({'error': 'Date and reason are required'}), 400
    conn = get_db()
    try:
        exists = conn.execute(
            "SELECT 1 FROM excuse_requests WHERE employee_id=? AND att_date=? AND vtype=? AND status='pending'",
            (emp_id, att_date, vtype)).fetchone()
        if exists:
            return jsonify({'error': 'A pending request already exists for this day'}), 400
        conn.execute(
            "INSERT INTO excuse_requests (employee_id, att_date, vtype, reason, attachment, attachment_name) VALUES (?,?,?,?,?,?)",
            (emp_id, att_date, vtype, reason, attachment, att_name))
        conn.commit()
        # notify managers
        _notify_excuse_submitted(emp_id, att_date, vtype, reason, conn)
        return jsonify({'ok': True, 'msg': 'Excuse submitted successfully'})
    finally:
        conn.close()

@app.route('/api/excuses/<int:eid>', methods=['PUT'])
@hr_required
def api_excuse_decide(eid):
    d      = request.get_json(silent=True) or {}
    status = d.get('status')
    note   = d.get('note', '')
    if status not in ('approved', 'rejected'):
        return jsonify({'error': 'Status must be approved or rejected'}), 400
    conn = get_db()
    try:
        conn.execute("""
            UPDATE excuse_requests
            SET status=?, decided_by=?, decided_at=datetime('now'), manager_note=?
            WHERE id=?
        """, (status, session['user_id'], note, eid))
        # if manager approved: delete associated violation
        if status == 'approved':
            ex = conn.execute(
                "SELECT * FROM excuse_requests WHERE id=?", (eid,)).fetchone()
            if ex:
                conn.execute("""
                    DELETE FROM violations
                    WHERE employee_id=? AND vio_date=? AND vtype=?
                """, (ex['employee_id'], ex['att_date'], ex['vtype']))
                # if absence and approved → update attendance status
                if ex['vtype'] == 'absent':
                    conn.execute("""
                        UPDATE attendance SET status='excused'
                        WHERE employee_id=? AND att_date=?
                    """, (ex['employee_id'], ex['att_date']))
        conn.commit()
        audit_log('excuse_decide', 'excuse', str(eid), details=f"status={status}")
        _notify_excuse_decision(eid, status, note, conn)
        return jsonify({'ok': True})
    finally:
        conn.close()

def _notify_excuse_submitted(emp_id, att_date, vtype, reason, conn):
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp: return
    name = emp['name_ar']
    vtype_ar = {'late': 'Late Arrival', 'early_leave': 'Early Departure', 'absent': 'Absence'}.get(vtype, vtype)
    managers = conn.execute(
        "SELECT u.*, e.email AS memail FROM users u LEFT JOIN employees e ON e.id=u.employee_id "
        "WHERE u.role IN ('hr','manager')").fetchall()
    for mgr in managers:
        to = mgr['memail'] or EMAIL_FROM
        if not to: continue
        subj = f"📋 New Excuse Request — {name} — {att_date}"
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#3b82f6;margin-bottom:6px">📋 New Excuse Request</h2>
          <p>Employee <b>{name}</b> submitted an excuse for <b>{vtype_ar}</b> on <b>{att_date}</b>.</p>
          <p><b>Reason:</b> {reason}</p>
          <a href="{SITE_URL}" style="display:inline-block;background:#3b82f6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:700;margin-top:10px">
            Review Request
          </a>
        </div>"""
        send_email(to, subj, body)

def _notify_excuse_decision(excuse_id, status, note, conn):
    ex  = conn.execute("SELECT * FROM excuse_requests WHERE id=?", (excuse_id,)).fetchone()
    if not ex: return
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (ex['employee_id'],)).fetchone()
    if not emp or not emp['email']: return
    name = emp['name_ar']
    status_ar = 'Approved ✅' if status == 'approved' else 'Rejected ❌'
    color = '#16a34a' if status == 'approved' else '#dc2626'
    vtype_ar = {'late': 'Late Arrival', 'early_leave': 'Early Departure', 'absent': 'Absence'}.get(ex['vtype'], ex['vtype'])
    note_row = f"<p><b>Manager Note:</b> {note}</p>" if note else ''
    subj = f"{'✅' if status=='approved' else '❌'} Excuse Decision — {ex['att_date']}"
    body = f"""<div style="{_STYLE}">
      <h2 style="color:{color};margin-bottom:6px">{status_ar} — {vtype_ar} Excuse</h2>
      <p>Dear <b>{name}</b>,</p>
      <p>Your excuse for <b>{vtype_ar}</b> on <b>{ex['att_date']}</b> has been <b style="color:{color}">{status_ar}</b>.</p>
      {note_row}
    </div>"""
    send_email(emp['email'], subj, body)

# ═══════════════════════════════════════════════════════════
#  LEAVES ROUTES
# ═══════════════════════════════════════════════════════════
LEAVE_NAMES = {
    'annual':   'Annual Leave',
    'sick':     'Sick Leave',
    'emergency':'Emergency Leave',
    'official': 'Official Holiday',
}

@app.route('/api/leaves', methods=['GET'])
@login_required
def api_leaves_get():
    leave_type = request.args.get('leave_type')
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        role   = session.get('role')
        emp_id = session.get('employee_id')
        base = "SELECT l.*, e.name_ar FROM leaves l JOIN employees e ON e.id=l.employee_id"
        if role in ('hr', 'manager'):
            where = "WHERE l.start_date LIKE ?"
            params = [prefix]
            if leave_type:
                where += " AND l.leave_type=?"
                params.append(leave_type)
            rows = conn.execute(f"{base} {where} ORDER BY l.created_at DESC", params).fetchall()
        else:
            if not emp_id: return jsonify([])
            where = "WHERE l.employee_id=? AND l.start_date LIKE ?"
            params = [emp_id, prefix]
            if leave_type:
                where += " AND l.leave_type=?"
                params.append(leave_type)
            rows = conn.execute(f"{base} {where} ORDER BY l.created_at DESC", params).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/leaves/timeline')
@hr_required
def api_leaves_timeline():
    from calendar import monthrange
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    days_in_month = monthrange(y, m)[1]
    month_start = str(date(y, m, 1))
    month_end   = str(date(y, m, days_in_month))
    conn = get_db()
    try:
        leaves = conn.execute("""
            SELECT l.employee_id, l.leave_type, l.start_date, l.end_date, l.days,
                   e.name_ar, e.emp_code
            FROM leaves l JOIN employees e ON e.id=l.employee_id
            WHERE l.status='approved' AND l.end_date>=? AND l.start_date<=?
            ORDER BY e.name_ar
        """, (month_start, month_end)).fetchall()
        emps = conn.execute(
            "SELECT id, name_ar, emp_code FROM employees ORDER BY name_ar").fetchall()
    finally:
        conn.close()

    from datetime import timedelta
    result = []
    for emp in emps:
        emp_leaves = [l for l in leaves if l['employee_id'] == emp['id']]
        if not emp_leaves:
            continue
        days_set = set()
        details  = []
        for l in emp_leaves:
            s = datetime.strptime(l['start_date'], '%Y-%m-%d').date()
            e2 = datetime.strptime(l['end_date'],   '%Y-%m-%d').date()
            cur = max(s, date(y, m, 1))
            while cur <= min(e2, date(y, m, days_in_month)):
                days_set.add(cur.day)
                cur += timedelta(days=1)
            details.append({'type': l['leave_type'], 'start': l['start_date'],
                            'end': l['end_date'], 'days': l['days']})
        result.append({'employee_id': emp['id'], 'name_ar': emp['name_ar'],
                       'emp_code': emp['emp_code'], 'leave_days': sorted(days_set),
                       'details': details})
    return jsonify({'rows': result, 'days_in_month': days_in_month, 'year': y, 'month': m})

@app.route('/api/leaves', methods=['POST'])
@login_required
def api_leaves_post():
    d        = request.get_json(silent=True) or {}
    role     = session.get('role')
    emp_id   = session.get('employee_id')
    if role in ('hr', 'manager'):
        emp_id = d.get('employee_id', emp_id)
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    leave_type = d.get('leave_type', '')
    start_date = d.get('start_date', '')
    end_date   = d.get('end_date', '')
    notes      = d.get('notes', '')
    if not leave_type or not start_date or not end_date:
        return jsonify({'error': 'Leave type and dates are required'}), 400
    try:
        s = date.fromisoformat(start_date)
        e_d = date.fromisoformat(end_date)
        days = (e_d - s).days + 1
        if days <= 0:
            return jsonify({'error': 'End date must be after start date'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    # sick leave: requires document — accept request but note it
    sick_doc   = d.get('sick_doc', '')
    attachment = d.get('attachment', '')
    att_name   = d.get('attachment_name', '')
    # HR approves directly, employee waits for approval
    init_status = 'approved' if role in ('hr', 'manager') else 'pending'
    approved_by = session['user_id'] if init_status == 'approved' else None

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO leaves
                (employee_id, leave_type, start_date, end_date, days,
                 status, approved_by, sick_doc, notes, attachment, attachment_name)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (emp_id, leave_type, start_date, end_date, days,
              init_status, approved_by, sick_doc, notes, attachment, att_name))
        conn.commit()
        return jsonify({'ok': True, 'msg': 'Leave registered', 'days': days})
    finally:
        conn.close()

@app.route('/api/leaves/<int:lid>', methods=['PUT'])
@hr_required
def api_leave_decide(lid):
    d = request.get_json(silent=True) or {}
    status = d.get('status')
    if status not in ('approved', 'rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db()
    try:
        conn.execute("""
            UPDATE leaves SET status=?, approved_by=?, notes=COALESCE(?,notes)
            WHERE id=?
        """, (status, session['user_id'], d.get('notes'), lid))
        conn.commit()
        audit_log('leave_decide', 'leave', str(lid), details=f"status={status}")
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/leaves/<int:lid>/attachment')
@login_required
def api_leave_attachment(lid):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment, attachment_name FROM leaves WHERE id=?", (lid,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row['attachment']:
        return jsonify({'error': 'No attachment found'}), 404
    return jsonify({'data': row['attachment'], 'name': row['attachment_name']})

@app.route('/api/leaves/<int:lid>', methods=['DELETE'])
@hr_required
def api_leave_delete(lid):
    conn = get_db()
    try:
        conn.execute("DELETE FROM leaves WHERE id=?", (lid,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  PUBLIC HOLIDAYS ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/api/holidays', methods=['GET'])
@login_required
def api_holidays_get():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM public_holidays ORDER BY h_date").fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/holidays', methods=['POST'])
@hr_required
def api_holidays_post():
    d    = request.get_json(silent=True) or {}
    hd   = d.get('h_date', '').strip()
    name = d.get('name', '').strip()
    if not hd or not name:
        return jsonify({'error': 'Date and name are required'}), 400
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO public_holidays (h_date, name, created_by) VALUES (?,?,?)",
            (hd, name, session['user_id']))
        conn.commit()
        return jsonify({'ok': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'This date already exists'}), 400
    finally:
        conn.close()

@app.route('/api/holidays/<int:hid>', methods=['DELETE'])
@hr_required
def api_holiday_delete(hid):
    conn = get_db()
    try:
        conn.execute("DELETE FROM public_holidays WHERE id=?", (hid,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  OVERTIME ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/api/overtime', methods=['GET'])
@login_required
def api_overtime_get():
    conn = get_db()
    try:
        role   = session.get('role')
        emp_id = session.get('employee_id')
        if role in ('hr', 'manager'):
            rows = conn.execute("""
                SELECT ot.*, e.name_ar FROM overtime_requests ot
                JOIN employees e ON e.id=ot.employee_id
                ORDER BY ot.att_date DESC
            """).fetchall()
        else:
            if not emp_id: return jsonify([])
            rows = conn.execute("""
                SELECT ot.*, e.name_ar FROM overtime_requests ot
                JOIN employees e ON e.id=ot.employee_id
                WHERE ot.employee_id=? ORDER BY ot.att_date DESC
            """, (emp_id,)).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/overtime', methods=['POST'])
@hr_required
def api_overtime_post():
    d          = request.get_json(silent=True) or {}
    emp_id     = d.get('employee_id')
    att_date   = d.get('att_date', '')
    ot_hours   = d.get('overtime_hours')
    notes      = (d.get('notes') or '').strip()
    if not emp_id or not att_date or not ot_hours:
        return jsonify({'error': 'Employee, date, and hours are required'}), 400
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO overtime_requests
                (employee_id, att_date, overtime_hours, check_out_time, work_end, status, notes, source)
            VALUES (?, ?, ?, '', '', 'approved', ?, 'manual')
        """, (emp_id, att_date, float(ot_hours), notes))
        conn.commit()
        emp_row = conn.execute("SELECT name_ar FROM employees WHERE id=?", (emp_id,)).fetchone()
        emp_name = emp_row['name_ar'] if emp_row else str(emp_id)
        audit_log('create_overtime', 'overtime', emp_name, details=f"date={att_date} hours={ot_hours} source=manual")
        return jsonify({'ok': True, 'msg': 'Assignment submitted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/overtime/<int:oid>', methods=['PUT'])
@hr_required
def api_overtime_decide(oid):
    d = request.get_json(silent=True) or {}
    status = d.get('status')
    note   = d.get('note', '')
    if status not in ('approved', 'rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db()
    try:
        conn.execute("""
            UPDATE overtime_requests
            SET status=?, decided_by=?, decided_at=datetime('now'), manager_note=?
            WHERE id=?
        """, (status, session['user_id'], note, oid))
        conn.commit()
        audit_log('overtime_decide', 'overtime', str(oid), details=f"status={status}")
    finally:
        conn.close()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  EMPLOYEE SELF-SERVICE ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/api/my/attendance')
@login_required
def api_my_attendance():
    emp_id = session.get('employee_id')
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT a.*, e.name_ar
            FROM attendance a JOIN employees e ON e.id=a.employee_id
            WHERE a.employee_id=? AND a.att_date LIKE ?
            ORDER BY a.att_date DESC
        """, (emp_id, prefix)).fetchall()
        # add excuse status to each attendance record
        result = []
        for r in rows:
            d = dict(r)
            ex = conn.execute(
                "SELECT status FROM excuse_requests WHERE employee_id=? AND att_date=?",
                (emp_id, r['att_date'])).fetchone()
            d['excuse_status'] = ex['status'] if ex else None
            result.append(d)
        leave_bal = _leave_balance(conn, emp_id, y)
    finally:
        conn.close()
    return jsonify({'records': result, 'leave_balance': leave_bal})

@app.route('/api/my/violations')
@login_required
def api_my_violations():
    emp_id = session.get('employee_id')
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT v.*, e.name_ar
            FROM violations v JOIN employees e ON e.id=v.employee_id
            WHERE v.employee_id=? AND v.vio_date LIKE ?
            ORDER BY v.vio_date DESC
        """, (emp_id, prefix)).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            ex = conn.execute(
                "SELECT id, status FROM excuse_requests WHERE employee_id=? AND att_date=? AND vtype=?",
                (emp_id, r['vio_date'], r['vtype'])).fetchone()
            d['excuse_id']     = ex['id']     if ex else None
            d['excuse_status'] = ex['status'] if ex else None
            result.append(d)
    finally:
        conn.close()
    return jsonify(result)

@app.route('/api/my/payroll')
@login_required
def api_my_payroll():
    emp_id = session.get('employee_id')
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    y = request.args.get('year',  date.today().year,  type=int)
    m = request.args.get('month', date.today().month, type=int)
    prefix = f"{y}-{m:02d}-%"
    conn = get_db()
    try:
        emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
        if not emp:
            return jsonify({'error': 'Employee not found'}), 404
        emp = dict(emp)
        vios = conn.execute(
            "SELECT * FROM violations WHERE employee_id=? AND vio_date LIKE ? ORDER BY vio_date",
            (emp_id, prefix)).fetchall()
        atts = conn.execute(
            "SELECT * FROM attendance WHERE employee_id=? AND att_date LIKE ? ORDER BY att_date",
            (emp_id, prefix)).fetchall()
        total_ded = sum(v['deduction'] for v in vios)
        gross    = (emp['salary'] or 0) + (emp['housing'] or 0) + (emp['transport'] or 0) + (emp['commission'] or 0)
        gosi_ded = _gosi(emp)
        net      = gross - total_ded - gosi_ded
        emp_out  = dict(emp)
        for k in ('salary','housing','transport','commission','other_ded'):
            emp_out[k] = emp_out.get(k) or 0
        return jsonify({
            'employee': emp_out,
            'year': y, 'month': m,
            'gross': round(gross, 2),
            'deductions': round(total_ded, 2),
            'other_ded': gosi_ded,
            'net': round(net, 2),
            'attendance_days': len([a for a in atts if a['status'] != 'absent']),
            'absent_days': len([a for a in atts if a['status'] == 'absent']),
            'late_days': len([a for a in atts if a['status'] == 'late']),
            'violations': [dict(v) for v in vios],
            'attendance': [dict(a) for a in atts],
        })
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════════
#  ATTENDANCE REQUESTS (late arrival / early departure requests)
# ═══════════════════════════════════════════════════════════
@app.route('/api/requests', methods=['GET'])
@login_required
def api_requests_get():
    conn = get_db()
    try:
        role   = session.get('role')
        emp_id = session.get('employee_id')
        if role in ('hr', 'manager'):
            rows = conn.execute("""
                SELECT ar.*, e.name_ar FROM attendance_requests ar
                JOIN employees e ON e.id=ar.employee_id
                ORDER BY ar.submitted_at DESC
            """).fetchall()
        else:
            if not emp_id: return jsonify([])
            rows = conn.execute("""
                SELECT ar.*, e.name_ar FROM attendance_requests ar
                JOIN employees e ON e.id=ar.employee_id
                WHERE ar.employee_id=? ORDER BY ar.submitted_at DESC
            """, (emp_id,)).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/requests', methods=['POST'])
@login_required
def api_requests_post():
    d       = request.get_json(silent=True) or {}
    role    = session.get('role')
    emp_id  = session.get('employee_id')
    if role in ('hr', 'manager'):
        emp_id = d.get('employee_id', emp_id)
    if not emp_id:
        return jsonify({'error': 'User is not linked to an employee'}), 400
    req_date       = d.get('req_date', '')
    req_type       = d.get('req_type', '')
    reason         = (d.get('reason') or '').strip()
    requested_time = d.get('requested_time', '')
    attachment     = d.get('attachment', '')
    att_name       = d.get('attachment_name', '')
    if not req_date or not req_type or not reason:
        return jsonify({'error': 'Date, type, and reason are required'}), 400
    if req_type not in ('late_arrival', 'early_leave'):
        return jsonify({'error': 'Invalid request type'}), 400
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO attendance_requests
                (employee_id, req_date, req_type, reason, requested_time,
                 attachment, attachment_name)
            VALUES (?,?,?,?,?,?,?)
        """, (emp_id, req_date, req_type, reason, requested_time, attachment, att_name))
        conn.commit()
        _notify_request_submitted(emp_id, req_date, req_type, reason, conn)
        return jsonify({'ok': True, 'msg': 'Request submitted successfully'})
    finally:
        conn.close()

@app.route('/api/requests/<int:rid>', methods=['PUT'])
@hr_required
def api_request_decide(rid):
    d    = request.get_json(silent=True) or {}
    status = d.get('status')
    note   = d.get('note', '')
    if status not in ('approved', 'rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db()
    try:
        conn.execute("""
            UPDATE attendance_requests
            SET status=?, decided_by=?, decided_at=datetime('now'), manager_note=?
            WHERE id=?
        """, (status, session['user_id'], note, rid))
        conn.commit()
        _notify_request_decision(rid, status, note, conn)
        return jsonify({'ok': True})
    finally:
        conn.close()

@app.route('/api/requests/<int:rid>/attachment')
@login_required
def api_request_attachment(rid):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment, attachment_name FROM attendance_requests WHERE id=?", (rid,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row['attachment']:
        return jsonify({'error': 'No attachment found'}), 404
    return jsonify({'data': row['attachment'], 'name': row['attachment_name']})

@app.route('/api/excuses/<int:eid>/attachment')
@login_required
def api_excuse_attachment(eid):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment, attachment_name FROM excuse_requests WHERE id=?", (eid,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row['attachment']:
        return jsonify({'error': 'No attachment found'}), 404
    return jsonify({'data': row['attachment'], 'name': row['attachment_name']})

def _notify_request_submitted(emp_id, req_date, req_type, reason, conn):
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp: return
    name = emp['name_ar']
    type_ar = 'Late Arrival' if req_type == 'late_arrival' else 'Early Departure'
    managers = conn.execute(
        "SELECT u.*, e.email AS memail FROM users u LEFT JOIN employees e ON e.id=u.employee_id "
        "WHERE u.role IN ('hr','manager')").fetchall()
    for mgr in managers:
        to = mgr['memail'] or EMAIL_FROM
        if not to: continue
        subj = f"📨 {type_ar} Request — {name} — {req_date}"
        body = f"""<div style="{_STYLE}">
          <h2 style="color:#8b5cf6;margin-bottom:6px">📨 {type_ar} Request</h2>
          <p>Employee <b>{name}</b> submitted a request on <b>{req_date}</b>.</p>
          <p><b>Reason:</b> {reason}</p>
          <a href="{SITE_URL}" style="display:inline-block;background:#8b5cf6;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:700;margin-top:10px">
            Review Request
          </a>
        </div>"""
        send_email(to, subj, body)

def _notify_request_decision(req_id, status, note, conn):
    req = conn.execute("SELECT * FROM attendance_requests WHERE id=?", (req_id,)).fetchone()
    if not req: return
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (req['employee_id'],)).fetchone()
    if not emp or not emp['email']: return
    type_ar  = 'Late Arrival' if req['req_type'] == 'late_arrival' else 'Early Departure'
    status_ar = 'Approved ✅' if status == 'approved' else 'Rejected ❌'
    color = '#16a34a' if status == 'approved' else '#dc2626'
    note_row = f"<p><b>Note:</b> {note}</p>" if note else ''
    subj = f"{'✅' if status=='approved' else '❌'} {type_ar} Request Decision — {req['req_date']}"
    body = f"""<div style="{_STYLE}">
      <h2 style="color:{color};margin-bottom:6px">{status_ar} — {type_ar} Request</h2>
      <p>Dear <b>{emp['name_ar']}</b>,</p>
      <p>Your request on <b>{req['req_date']}</b> has been <b style="color:{color}">{status_ar}</b>.</p>
      {note_row}
    </div>"""
    send_email(emp['email'], subj, body)

# ═══════════════════════════════════════════════════════════
#  AUTO-REJECT EXCUSES JOB
# ═══════════════════════════════════════════════════════════
def auto_reject_excuses():
    """Auto-reject pending excuses and requests after AUTO_REJECT_DAYS days"""
    conn = get_db()
    try:
        cutoff = str(datetime.now() - timedelta(days=AUTO_REJECT_DAYS))
        note   = f'Auto-rejected — no response within {AUTO_REJECT_DAYS} days'

        # excuses
        excuses = conn.execute(
            "SELECT * FROM excuse_requests WHERE status='pending' AND submitted_at<?",
            (cutoff,)).fetchall()
        for ex in excuses:
            conn.execute(
                "UPDATE excuse_requests SET status='rejected', decided_at=datetime('now'), manager_note=? WHERE id=?",
                (note, ex['id']))
            _notify_excuse_decision(ex['id'], 'rejected', note, conn)

        # attendance requests
        reqs = conn.execute(
            "SELECT * FROM attendance_requests WHERE status='pending' AND submitted_at<?",
            (cutoff,)).fetchall()
        for rq in reqs:
            conn.execute(
                "UPDATE attendance_requests SET status='rejected', decided_at=datetime('now'), manager_note=? WHERE id=?",
                (note, rq['id']))
            _notify_request_decision(rq['id'], 'rejected', note, conn)

        # leaves
        leaves = conn.execute(
            "SELECT * FROM leaves WHERE status='pending' AND created_at<?",
            (cutoff,)).fetchall()
        for lv in leaves:
            conn.execute(
                "UPDATE leaves SET status='rejected', approved_by=NULL, notes=? WHERE id=?",
                (note, lv['id']))

        conn.commit()
        total = len(excuses) + len(reqs)
        if total:
            logger.info(f"Auto-rejected {total} pending requests")
    except Exception as e:
        logger.error(f"auto_reject_excuses error: {e}")
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════════
#  SCHEDULER
# ═══════════════════════════════════════════════════════════
def scheduled_job():
    logger.info("Scheduler triggered — running daily attendance processing")
    process_day()

# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == '__main__':
    init_db()
    scheduler = BackgroundScheduler(timezone='Asia/Riyadh')
    scheduler.add_job(scheduled_job, CronTrigger(hour=20, minute=0))
    scheduler.add_job(auto_reject_excuses, CronTrigger(hour=8, minute=0))
    scheduler.start()
    logger.info("Scheduler started — attendance 20:00, auto-reject 08:00 AST")
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
